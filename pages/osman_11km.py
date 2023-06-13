import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile
from PIL import Image

# menghilangkan hamburger
st.markdown("""
<style>
.css-1rs6os.edgvbvh3
{
    visibility:hidden;
}
.css-1lsmgbg.egzxvld0
{
    visibility:hidden;
}
</style>
""", unsafe_allow_html=True)

image = Image.open('logo resmi nf resize.png')
st.image(image)

st.title("Olah Nilai Standar")
st.header("11 SMA KM")

col6 = st.container()

with col6:
    KELAS = st.selectbox(
        "KELAS",
        ("--Pilih Kelas--", "11 SMA"))

col7 = st.container()

with col7:
    SEMESTER = st.selectbox(
        "SEMESTER",
        ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

col8 = st.container()

with col8:
    PENILAIAN = st.selectbox(
        "PENILAIAN",
        ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN"))

col9 = st.container()

with col9:
    KURIKULUM = st.selectbox(
        "KURIKULUM",
        ("--Pilih Kurikulum--", "KM"))

TAHUN = st.text_input("Masukkan Tahun Ajaran", placeholder="contoh: 2022-2023")

col1, col2, col3, col4, col5, col6 = st.columns(
    6)

with col1:
    MTW = st.selectbox(
        "JML. SOAL MAW.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col2:
    MTP = st.selectbox(
        "JML. SOAL MAP.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col3:
    IND = st.selectbox(
        "JML. SOAL IND.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col4:
    ENG = st.selectbox(
        "JML. SOAL ENG.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col5:
    SEJ = st.selectbox(
        "JML. SOAL SEJ.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col6:
    GEO = st.selectbox(
        "JML. SOAL GEO.",
        (15, 20, 25, 30, 35, 40, 45, 50))

col7, col8, col9, col10, col11 = st.columns(5)

with col7:
    EKO = st.selectbox(
        "JML. SOAL EKO.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col8:
    SOS = st.selectbox(
        "JML. SOAL SOS.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col9:
    FIS = st.selectbox(
        "JML. SOAL FIS.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col10:
    KIM = st.selectbox(
        "JML. SOAL KIM.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col11:
    BIO = st.selectbox(
        "JML. SOAL BIO.",
        (15, 20, 25, 30, 35, 40, 45, 50))


JML_SOAL_MAW = MTW
JML_SOAL_MAP = MTP
JML_SOAL_IND = IND
JML_SOAL_ENG = ENG
JML_SOAL_SEJ = SEJ
JML_SOAL_GEO = GEO
JML_SOAL_EKO = EKO
JML_SOAL_SOS = SOS
JML_SOAL_FIS = FIS
JML_SOAL_KIM = KIM
JML_SOAL_BIO = BIO

uploaded_file = st.file_uploader('Letakkan file excel 11 SMA', type='xlsx')

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb['Sheet1']

    q = len(ws['K'])
    r = len(ws['K'])+2
    s = len(ws['K'])+3
    t = len(ws['K'])+4
    u = len(ws['K'])+5
    v = len(ws['K'])+6
    w = len(ws['K'])+7
    x = len(ws['K'])+8

    ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
    ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
    ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
    ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
    ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
    ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)
    ws['M{}'.format(r)] = "=ROUND(AVERAGE(M2:M{}),2)".format(q)
    ws['N{}'.format(r)] = "=ROUND(AVERAGE(N2:N{}),2)".format(q)
    ws['O{}'.format(r)] = "=ROUND(AVERAGE(O2:O{}),2)".format(q)
    ws['P{}'.format(r)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
    ws['Q{}'.format(r)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
    ws['R{}'.format(r)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)

    ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
    ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
    ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
    ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
    ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
    ws['L{}'.format(s)] = "=STDEV(L2:L{})".format(q)
    ws['M{}'.format(s)] = "=STDEV(M2:M{})".format(q)
    ws['N{}'.format(s)] = "=STDEV(N2:N{})".format(q)
    ws['O{}'.format(s)] = "=STDEV(O2:O{})".format(q)
    ws['P{}'.format(s)] = "=STDEV(P2:P{})".format(q)
    ws['Q{}'.format(s)] = "=STDEV(Q2:Q{})".format(q)

    ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
    ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
    ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
    ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
    ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
    ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
    ws['M{}'.format(t)] = "=MAX(M2:M{})".format(q)
    ws['N{}'.format(t)] = "=MAX(N2:N{})".format(q)
    ws['O{}'.format(t)] = "=MAX(O2:O{})".format(q)
    ws['P{}'.format(t)] = "=MAX(P2:P{})".format(q)
    ws['Q{}'.format(t)] = "=MAX(Q2:Q{})".format(q)

    ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
    ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
    ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
    ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
    ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
    ws['W{}'.format(r)] = "=MAX(W2:W{})".format(q)
    ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
    ws['Y{}'.format(r)] = "=MAX(Y2:Y{})".format(q)
    ws['Z{}'.format(r)] = "=MAX(Z2:Z{})".format(q)
    ws['AA{}'.format(r)] = "=MAX(AA2:AA{})".format(q)
    ws['AB{}'.format(r)] = "=MAX(AB2:AB{})".format(q)
    ws['AC{}'.format(r)] = "=MAX(AC2:AC{})".format(q)
    ws['AD{}'.format(r)] = "=MAX(AD2:AD{})".format(q)
    ws['AE{}'.format(r)] = "=MAX(AE2:AE{})".format(q)
    ws['AF{}'.format(r)] = "=MAX(AF2:AF{})".format(q)
    ws['AG{}'.format(r)] = "=MAX(AG2:AG{})".format(q)
    ws['AH{}'.format(r)] = "=MAX(AH2:AH{})".format(q)
    ws['AI{}'.format(r)] = "=MAX(AI2:AI{})".format(q)
    ws['AJ{}'.format(r)] = "=MAX(AJ2:AJ{})".format(q)
    ws['AK{}'.format(r)] = "=MAX(AK2:AK{})".format(q)
    ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
    ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
    ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
    ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)
    ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)

    ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
    ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
    ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
    ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
    ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
    ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
    ws['M{}'.format(u)] = "=MIN(M2:M{})".format(q)
    ws['N{}'.format(u)] = "=MIN(N2:N{})".format(q)
    ws['O{}'.format(u)] = "=MIN(O2:O{})".format(q)
    ws['P{}'.format(u)] = "=MIN(P2:P{})".format(q)
    ws['Q{}'.format(u)] = "=MIN(Q2:Q{})".format(q)
    ws['R{}'.format(u)] = "=MIN(R2:R{})".format(q)

    ws['AD{}'.format(s)] = "=MIN(AD2:AD{})".format(q)
    ws['AE{}'.format(s)] = "=MIN(AE2:AE{})".format(q)
    ws['AF{}'.format(s)] = "=MIN(AF2:AF{})".format(q)
    ws['AG{}'.format(s)] = "=MIN(AG2:AG{})".format(q)
    ws['AH{}'.format(s)] = "=MIN(AH2:AH{})".format(q)
    ws['AI{}'.format(s)] = "=MIN(AI2:AI{})".format(q)
    ws['AJ{}'.format(s)] = "=MIN(AJ2:AJ{})".format(q)
    ws['AK{}'.format(s)] = "=MIN(AK2:AK{})".format(q)
    ws['AL{}'.format(s)] = "=MIN(AL2:AL{})".format(q)
    ws['AM{}'.format(s)] = "=MIN(AM2:AM{})".format(q)
    ws['AN{}'.format(s)] = "=MIN(AN2:AN{})".format(q)
    ws['AO{}'.format(s)] = "=MIN(AO2:AO{})".format(q)

    ws['AD{}'.format(t)] = "=ROUND(AVERAGE(AD2:AD{}),2)".format(q)
    ws['AE{}'.format(t)] = "=ROUND(AVERAGE(AE2:AE{}),2)".format(q)
    ws['AF{}'.format(t)] = "=ROUND(AVERAGE(AF2:AF{}),2)".format(q)
    ws['AG{}'.format(t)] = "=ROUND(AVERAGE(AG2:AG{}),2)".format(q)
    ws['AH{}'.format(t)] = "=ROUND(AVERAGE(AH2:AH{}),2)".format(q)
    ws['AI{}'.format(t)] = "=ROUND(AVERAGE(AI2:AI{}),2)".format(q)
    ws['AJ{}'.format(t)] = "=ROUND(AVERAGE(AJ2:AJ{}),2)".format(q)
    ws['AK{}'.format(t)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
    ws['AL{}'.format(t)] = "=ROUND(AVERAGE(AL2:AL{}),2)".format(q)
    ws['AM{}'.format(t)] = "=ROUND(AVERAGE(AM2:AM{}),2)".format(q)
    ws['AN{}'.format(t)] = "=ROUND(AVERAGE(AN2:AN{}),2)".format(q)
    ws['AO{}'.format(t)] = "=ROUND(AVERAGE(AO2:AO{}),2)".format(q)

    ws['AR{}'.format(r)] = "=SUM(AR2:AR{})".format(q)
    ws['AS{}'.format(r)] = "=SUM(AS2:AS{})".format(q)
    ws['AT{}'.format(r)] = "=SUM(AT2:AT{})".format(q)
    ws['AU{}'.format(r)] = "=SUM(AU2:AU{})".format(q)
    ws['AV{}'.format(r)] = "=SUM(AV2:AV{})".format(q)
    ws['AW{}'.format(r)] = "=SUM(AW2:AW{})".format(q)
    ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
    ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
    ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)
    ws['BA{}'.format(r)] = "=SUM(BA2:BA{})".format(q)
    ws['BB{}'.format(r)] = "=SUM(BB2:BB{})".format(q)

    # Jumlah Soal
    ws['F{}'.format(v)] = 'JUMLAH SOAL'
    ws['G{}'.format(v)] = JML_SOAL_MAW
    ws['H{}'.format(v)] = JML_SOAL_MAP
    ws['I{}'.format(v)] = JML_SOAL_IND
    ws['J{}'.format(v)] = JML_SOAL_ENG
    ws['K{}'.format(v)] = JML_SOAL_SEJ
    ws['L{}'.format(v)] = JML_SOAL_GEO
    ws['M{}'.format(v)] = JML_SOAL_EKO
    ws['N{}'.format(v)] = JML_SOAL_SOS
    ws['O{}'.format(v)] = JML_SOAL_FIS
    ws['P{}'.format(v)] = JML_SOAL_KIM
    ws['Q{}'.format(v)] = JML_SOAL_BIO

    # new
    # iterasi 1 rata-rata - 1
    # rata" MTW ke MTW tambahan dan mapel MTW awal
    ws['BI{}'.format(r)] = "=IF($AR${}=0,$G${},$G${}-1)".format(r, r, r)
    ws['BI{}'.format(s)] = "=STDEV(BI2:BI{})".format(q)
    ws['BI{}'.format(t)] = "=MAX(BI2:BI{})".format(q)
    ws['BI{}'.format(u)] = "=MIN(BI2:BI{})".format(q)
    # rata" MTP ke MTP tambahan dan mapel MTP awal
    ws['BJ{}'.format(r)] = "=IF($AS${}=0,$H${},$H${}-1)".format(r, r, r)
    ws['BJ{}'.format(s)] = "=STDEV(BJ2:BJ{})".format(q)
    ws['BJ{}'.format(t)] = "=MAX(BJ2:BJ{})".format(q)
    ws['BJ{}'.format(u)] = "=MIN(BJ2:BJ{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['BK{}'.format(r)] = "=IF($AT${}=0,$I${},$I${}-1)".format(r, r, r)
    ws['BK{}'.format(s)] = "=STDEV(BK2:BK{})".format(q)
    ws['BK{}'.format(t)] = "=MAX(BK2:BK{})".format(q)
    ws['BK{}'.format(u)] = "=MIN(BK2:BK{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['BL{}'.format(r)] = "=IF($AU${}=0,$J${},$J${}-1)".format(r, r, r)
    ws['BL{}'.format(s)] = "=STDEV(BL2:BL{})".format(q)
    ws['BL{}'.format(t)] = "=MAX(BL2:BL{})".format(q)
    ws['BL{}'.format(u)] = "=MIN(BL2:BL{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['BM{}'.format(r)] = "=IF($AV${}=0,$K${},$K${}-1)".format(r, r, r)
    ws['BM{}'.format(s)] = "=STDEV(BM2:BM{})".format(q)
    ws['BM{}'.format(t)] = "=MAX(BM2:BM{})".format(q)
    ws['BM{}'.format(u)] = "=MIN(BM2:BM{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['BN{}'.format(r)] = "=IF($AW${}=0,$L${},$L${}-1)".format(r, r, r)
    ws['BN{}'.format(s)] = "=STDEV(BN2:BN{})".format(q)
    ws['BN{}'.format(t)] = "=MAX(BN2:BN{})".format(q)
    ws['BN{}'.format(u)] = "=MIN(BN2:BN{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['BO{}'.format(r)] = "=IF($AX${}=0,$M${},$M${}-1)".format(r, r, r)
    ws['BO{}'.format(s)] = "=STDEV(BO2:BO{})".format(q)
    ws['BO{}'.format(t)] = "=MAX(BO2:BO{})".format(q)
    ws['BO{}'.format(u)] = "=MIN(BO2:BO{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['BP{}'.format(r)] = "=IF($AY${}=0,$N${},$N${}-1)".format(r, r, r)
    ws['BP{}'.format(s)] = "=STDEV(BP2:BP{})".format(q)
    ws['BP{}'.format(t)] = "=MAX(BP2:BP{})".format(q)
    ws['BP{}'.format(u)] = "=MIN(BP2:BP{})".format(q)
    # rata" FIS ke FIS tambahan dan mapel FIS awal
    ws['BQ{}'.format(r)] = "=IF($AZ${}=0,$O${},$O${}-1)".format(r, r, r)
    ws['BQ{}'.format(s)] = "=STDEV(BQ2:BQ{})".format(q)
    ws['BQ{}'.format(t)] = "=MAX(BQ2:BQ{})".format(q)
    ws['BQ{}'.format(u)] = "=MIN(BQ2:BQ{})".format(q)
    # rata" KIM ke KIM tambahan dan mapel KIM awal
    ws['BR{}'.format(r)] = "=IF($BA${}=0,$P${},$P${}-1)".format(r, r, r)
    ws['BR{}'.format(s)] = "=STDEV(BR2:BR{})".format(q)
    ws['BR{}'.format(t)] = "=MAX(BR2:BR{})".format(q)
    ws['BR{}'.format(u)] = "=MIN(BR2:BR{})".format(q)
    # rata" BIO ke BIO tambahan dan mapel BIO awal
    ws['BS{}'.format(r)] = "=IF($BB${}=0,$Q${},$Q${}-1)".format(r, r, r)
    ws['BS{}'.format(s)] = "=STDEV(BS2:BS{})".format(q)
    ws['BS{}'.format(t)] = "=MAX(BS2:BS{})".format(q)
    ws['BS{}'.format(u)] = "=MIN(BS2:BS{})".format(q)
    # jml MAPEL
    ws['BT{}'.format(r)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)
    ws['BT{}'.format(t)] = "=MAX(BT2:BT{})".format(q)
    ws['BT{}'.format(u)] = "=MIN(BT2:BT{})".format(q)
    # MAX Z SCORE
    ws['BU{}'.format(r)] = "=MAX(BU2:BU{})".format(q)
    ws['BV{}'.format(r)] = "=MAX(BV2:BV{})".format(q)
    ws['BW{}'.format(r)] = "=MAX(BW2:BW{})".format(q)
    ws['BX{}'.format(r)] = "=MAX(BX2:BX{})".format(q)
    ws['BY{}'.format(r)] = "=MAX(BY2:BY{})".format(q)
    ws['BZ{}'.format(r)] = "=MAX(BZ2:BZ{})".format(q)
    ws['CA{}'.format(r)] = "=MAX(CA2:CA{})".format(q)
    ws['CB{}'.format(r)] = "=MAX(CB2:CB{})".format(q)
    ws['CC{}'.format(r)] = "=MAX(CC2:CC{})".format(q)
    ws['CD{}'.format(r)] = "=MAX(CD2:CD{})".format(q)
    ws['CE{}'.format(r)] = "=MAX(CE2:CE{})".format(q)
    # NILAI STANDAR MTW
    ws['CF{}'.format(r)] = "=MAX(CF2:CF{})".format(q)
    ws['CF{}'.format(s)] = "=MIN(CF2:CF{})".format(q)
    ws['CF{}'.format(t)] = "=ROUND(AVERAGE(CF2:CF{}),2)".format(q)
    # NILAI STANDAR MTP
    ws['CG{}'.format(r)] = "=MAX(CG2:CG{})".format(q)
    ws['CG{}'.format(s)] = "=MIN(CG2:CG{})".format(q)
    ws['CG{}'.format(t)] = "=ROUND(AVERAGE(CG2:CG{}),2)".format(q)
    # NILAI STANDAR IND
    ws['CH{}'.format(r)] = "=MAX(CH2:CH{})".format(q)
    ws['CH{}'.format(s)] = "=MIN(CH2:CH{})".format(q)
    ws['CH{}'.format(t)] = "=ROUND(AVERAGE(CH2:CH{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['CI{}'.format(r)] = "=MAX(CI2:CI{})".format(q)
    ws['CI{}'.format(s)] = "=MIN(CI2:CI{})".format(q)
    ws['CI{}'.format(t)] = "=ROUND(AVERAGE(CI2:CI{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['CJ{}'.format(r)] = "=MAX(CJ2:CJ{})".format(q)
    ws['CJ{}'.format(s)] = "=MIN(CJ2:CJ{})".format(q)
    ws['CJ{}'.format(t)] = "=ROUND(AVERAGE(CJ2:CJ{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['CK{}'.format(r)] = "=MAX(CK2:CK{})".format(q)
    ws['CK{}'.format(s)] = "=MIN(CK2:CK{})".format(q)
    ws['CK{}'.format(t)] = "=ROUND(AVERAGE(CK2:CK{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
    ws['CL{}'.format(s)] = "=MIN(CL2:CL{})".format(q)
    ws['CL{}'.format(t)] = "=ROUND(AVERAGE(CL2:CL{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
    ws['CM{}'.format(s)] = "=MIN(CM2:CM{})".format(q)
    ws['CM{}'.format(t)] = "=ROUND(AVERAGE(CM2:CM{}),2)".format(q)
    # NILAI STANDAR FIS
    ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
    ws['CN{}'.format(s)] = "=MIN(CN2:CN{})".format(q)
    ws['CN{}'.format(t)] = "=ROUND(AVERAGE(CN2:CN{}),2)".format(q)
    # NILAI STANDAR KIM
    ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)
    ws['CO{}'.format(s)] = "=MIN(CO2:CO{})".format(q)
    ws['CO{}'.format(t)] = "=ROUND(AVERAGE(CO2:CO{}),2)".format(q)
    # NILAI STANDAR BIO
    ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
    ws['CP{}'.format(s)] = "=MIN(CP2:CP{})".format(q)
    ws['CP{}'.format(t)] = "=ROUND(AVERAGE(CP2:CP{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
    ws['CQ{}'.format(s)] = "=MIN(CQ2:CQ{})".format(q)
    ws['CQ{}'.format(t)] = "=ROUND(AVERAGE(CQ2:CQ{}),2)".format(q)

    # TAMBAHAN
    ws['CT{}'.format(r)] = "=SUM(CT2:CT{})".format(q)
    ws['CU{}'.format(r)] = "=SUM(CU2:CU{})".format(q)
    ws['CV{}'.format(r)] = "=SUM(CV2:CV{})".format(q)
    ws['CW{}'.format(r)] = "=SUM(CW2:CW{})".format(q)
    ws['CX{}'.format(r)] = "=SUM(CX2:CX{})".format(q)
    ws['CY{}'.format(r)] = "=SUM(CY2:CY{})".format(q)
    ws['CZ{}'.format(r)] = "=SUM(CZ2:CZ{})".format(q)
    ws['DA{}'.format(r)] = "=SUM(DA2:DA{})".format(q)
    ws['DB{}'.format(r)] = "=SUM(DB2:DB{})".format(q)
    ws['DC{}'.format(r)] = "=SUM(DC2:DC{})".format(q)
    ws['DD{}'.format(r)] = "=SUM(DD2:DD{})".format(q)

    # iterasi 2 rata-rata - 2
    # rata" MTW ke MTW tambahan dan mapel MTW awal
    ws['DK{}'.format(r)] = "=IF($CT${}=0,$BI${},$BI${}-1)".format(r, r, r)
    ws['DK{}'.format(s)] = "=STDEV(DK2:DK{})".format(q)
    ws['DK{}'.format(t)] = "=MAX(DK2:DK{})".format(q)
    ws['DK{}'.format(u)] = "=MIN(DK2:DK{})".format(q)
    # rata" MTP ke MTP tambahan dan mapel MTP awal
    ws['DL{}'.format(r)] = "=IF($CU${}=0,$BJ${},$BJ${}-1)".format(r, r, r)
    ws['DL{}'.format(s)] = "=STDEV(DL2:DL{})".format(q)
    ws['DL{}'.format(t)] = "=MAX(DL2:DL{})".format(q)
    ws['DL{}'.format(u)] = "=MIN(DL2:DL{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['DM{}'.format(r)] = "=IF($CV${}=0,$BK${},$BK${}-1)".format(r, r, r)
    ws['DM{}'.format(s)] = "=STDEV(DM2:DM{})".format(q)
    ws['DM{}'.format(t)] = "=MAX(DM2:DM{})".format(q)
    ws['DM{}'.format(u)] = "=MIN(DM2:DM{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['DN{}'.format(r)] = "=IF($CW${}=0,$BL${},$BL${}-1)".format(r, r, r)
    ws['DN{}'.format(s)] = "=STDEV(DN2:DN{})".format(q)
    ws['DN{}'.format(t)] = "=MAX(DN2:DN{})".format(q)
    ws['DN{}'.format(u)] = "=MIN(DN2:DN{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['DO{}'.format(r)] = "=IF($CX${}=0,$BM${},$BM${}-1)".format(r, r, r)
    ws['DO{}'.format(s)] = "=STDEV(DO2:DO{})".format(q)
    ws['DO{}'.format(t)] = "=MAX(DO2:DO{})".format(q)
    ws['DO{}'.format(u)] = "=MIN(DO2:DO{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['DP{}'.format(r)] = "=IF($CY${}=0,$BN${},$BN${}-1)".format(r, r, r)
    ws['DP{}'.format(s)] = "=STDEV(DP2:DP{})".format(q)
    ws['DP{}'.format(t)] = "=MAX(DP2:DP{})".format(q)
    ws['DP{}'.format(u)] = "=MIN(DP2:DP{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['DQ{}'.format(r)] = "=IF($CZ${}=0,$BO${},$BO${}-1)".format(r, r, r)
    ws['DQ{}'.format(s)] = "=STDEV(DQ2:DQ{})".format(q)
    ws['DQ{}'.format(t)] = "=MAX(DQ2:DQ{})".format(q)
    ws['DQ{}'.format(u)] = "=MIN(DQ2:DQ{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['DR{}'.format(r)] = "=IF($DA${}=0,$BP${},$BP${}-1)".format(r, r, r)
    ws['DR{}'.format(s)] = "=STDEV(DR2:DR{})".format(q)
    ws['DR{}'.format(t)] = "=MAX(DR2:DR{})".format(q)
    ws['DR{}'.format(u)] = "=MIN(DR2:DR{})".format(q)
    # rata" FIS ke FIS tambahan dan mapel FIS awal
    ws['DS{}'.format(r)] = "=IF($DB${}=0,$BQ${},$BQ${}-1)".format(r, r, r)
    ws['DS{}'.format(s)] = "=STDEV(DS2:DS{})".format(q)
    ws['DS{}'.format(t)] = "=MAX(DS2:DS{})".format(q)
    ws['DS{}'.format(u)] = "=MIN(DS2:DS{})".format(q)
    # rata" KIM ke KIM tambahan dan mapel KIM awal
    ws['DT{}'.format(r)] = "=IF($DC${}=0,$BR${},$BR${}-1)".format(r, r, r)
    ws['DT{}'.format(s)] = "=STDEV(DT2:DT{})".format(q)
    ws['DT{}'.format(t)] = "=MAX(DT2:DT{})".format(q)
    ws['DT{}'.format(u)] = "=MIN(DT2:DT{})".format(q)
    # rata" BIO ke BIO tambahan dan mapel BIO awal
    ws['DU{}'.format(r)] = "=IF($DD${}=0,$BS${},$BS${}-1)".format(r, r, r)
    ws['DU{}'.format(s)] = "=STDEV(DU2:DU{})".format(q)
    ws['DU{}'.format(t)] = "=MAX(DU2:DU{})".format(q)
    ws['DU{}'.format(u)] = "=MIN(DU2:DU{})".format(q)
    # jml MAPEL
    ws['DV{}'.format(r)] = "=ROUND(AVERAGE(DV2:DV{}),2)".format(q)
    ws['DV{}'.format(t)] = "=MAX(DV2:DV{})".format(q)
    ws['DV{}'.format(u)] = "=MIN(DV2:DV{})".format(q)
    # MAX Z SCORE
    ws['DW{}'.format(r)] = "=MAX(DW2:DW{})".format(q)
    ws['DX{}'.format(r)] = "=MAX(DX2:DX{})".format(q)
    ws['DY{}'.format(r)] = "=MAX(DY2:DY{})".format(q)
    ws['DZ{}'.format(r)] = "=MAX(DZ2:DZ{})".format(q)
    ws['EA{}'.format(r)] = "=MAX(EA2:EA{})".format(q)
    ws['EB{}'.format(r)] = "=MAX(EB2:EB{})".format(q)
    ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
    ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
    ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
    ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
    ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
    # NILAI STANDAR MTW
    ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
    ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
    ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
    # NILAI STANDAR MTP
    ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
    ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
    ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
    # NILAI STANDAR IND
    ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
    ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
    ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
    ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
    ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
    ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
    ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
    ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
    ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['EN{}'.format(r)] = "=MAX(EN2:EN{})".format(q)
    ws['EN{}'.format(s)] = "=MIN(EN2:EN{})".format(q)
    ws['EN{}'.format(t)] = "=ROUND(AVERAGE(EN2:EN{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)
    ws['EO{}'.format(s)] = "=MIN(EO2:EO{})".format(q)
    ws['EO{}'.format(t)] = "=ROUND(AVERAGE(EO2:EO{}),2)".format(q)
    # NILAI STANDAR FIS
    ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
    ws['EP{}'.format(s)] = "=MIN(EP2:EP{})".format(q)
    ws['EP{}'.format(t)] = "=ROUND(AVERAGE(EP2:EP{}),2)".format(q)
    # NILAI STANDAR KIM
    ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
    ws['EQ{}'.format(s)] = "=MIN(EQ2:EQ{})".format(q)
    ws['EQ{}'.format(t)] = "=ROUND(AVERAGE(EQ2:EQ{}),2)".format(q)
    # NILAI STANDAR BIO
    ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
    ws['ER{}'.format(s)] = "=MIN(ER2:ER{})".format(q)
    ws['ER{}'.format(t)] = "=ROUND(AVERAGE(ER2:ER{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
    ws['ES{}'.format(s)] = "=MIN(ES2:ES{})".format(q)
    ws['ES{}'.format(t)] = "=ROUND(AVERAGE(ES2:ES{}),2)".format(q)

    # TAMBAHAN
    ws['EV{}'.format(r)] = "=SUM(EV2:EV{})".format(q)
    ws['EW{}'.format(r)] = "=SUM(EW2:EW{})".format(q)
    ws['EX{}'.format(r)] = "=SUM(EX2:EX{})".format(q)
    ws['EY{}'.format(r)] = "=SUM(EY2:EY{})".format(q)
    ws['EZ{}'.format(r)] = "=SUM(EZ2:EZ{})".format(q)
    ws['FA{}'.format(r)] = "=SUM(FA2:FA{})".format(q)
    ws['FB{}'.format(r)] = "=SUM(FB2:FB{})".format(q)
    ws['FC{}'.format(r)] = "=SUM(FC2:FC{})".format(q)
    ws['FD{}'.format(r)] = "=SUM(FD2:FD{})".format(q)
    ws['FE{}'.format(r)] = "=SUM(FE2:FE{})".format(q)
    ws['FF{}'.format(r)] = "=SUM(FF2:FF{})".format(q)

    # iterasi 3 rata-rata - 3
    # rata" MTW ke MTW tambahan dan mapel MTW awal
    ws['FM{}'.format(r)] = "=IF($EV${}=0,$DK${},$DK${}-1)".format(r, r, r)
    ws['FM{}'.format(s)] = "=STDEV(FM2:FM{})".format(q)
    ws['FM{}'.format(t)] = "=MAX(FM2:FM{})".format(q)
    ws['FM{}'.format(u)] = "=MIN(FM2:FM{})".format(q)
    # rata" MTP ke MTP tambahan dan mapel MTP awal
    ws['FN{}'.format(r)] = "=IF($EW${}=0,$DL${},$DL${}-1)".format(r, r, r)
    ws['FN{}'.format(s)] = "=STDEV(FN2:FN{})".format(q)
    ws['FN{}'.format(t)] = "=MAX(FN2:FN{})".format(q)
    ws['FN{}'.format(u)] = "=MIN(FN2:FN{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['FO{}'.format(r)] = "=IF($EX${}=0,$DM${},$DM${}-1)".format(r, r, r)
    ws['FO{}'.format(s)] = "=STDEV(FO2:FO{})".format(q)
    ws['FO{}'.format(t)] = "=MAX(FO2:FO{})".format(q)
    ws['FO{}'.format(u)] = "=MIN(FO2:FO{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['FP{}'.format(r)] = "=IF($EY${}=0,$DN${},$DN${}-1)".format(r, r, r)
    ws['FP{}'.format(s)] = "=STDEV(FP2:FP{})".format(q)
    ws['FP{}'.format(t)] = "=MAX(FP2:FP{})".format(q)
    ws['FP{}'.format(u)] = "=MIN(FP2:FP{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['FQ{}'.format(r)] = "=IF($EZ${}=0,$DO${},$DO${}-1)".format(r, r, r)
    ws['FQ{}'.format(s)] = "=STDEV(FQ2:FQ{})".format(q)
    ws['FQ{}'.format(t)] = "=MAX(FQ2:FQ{})".format(q)
    ws['FQ{}'.format(u)] = "=MIN(FQ2:FQ{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['FR{}'.format(r)] = "=IF($FA${}=0,$DP${},$DP${}-1)".format(r, r, r)
    ws['FR{}'.format(s)] = "=STDEV(FR2:FR{})".format(q)
    ws['FR{}'.format(t)] = "=MAX(FR2:FR{})".format(q)
    ws['FR{}'.format(u)] = "=MIN(FR2:FR{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['FS{}'.format(r)] = "=IF($FB${}=0,$DQ${},$DQ${}-1)".format(r, r, r)
    ws['FS{}'.format(s)] = "=STDEV(FS2:FS{})".format(q)
    ws['FS{}'.format(t)] = "=MAX(FS2:FS{})".format(q)
    ws['FS{}'.format(u)] = "=MIN(FS2:FS{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['FT{}'.format(r)] = "=IF($FC${}=0,$DR${},$DR${}-1)".format(r, r, r)
    ws['FT{}'.format(s)] = "=STDEV(FT2:FT{})".format(q)
    ws['FT{}'.format(t)] = "=MAX(FT2:FT{})".format(q)
    ws['FT{}'.format(u)] = "=MIN(FT2:FT{})".format(q)
    # rata" FIS ke FIS tambahan dan mapel FIS awal
    ws['FU{}'.format(r)] = "=IF($FD${}=0,$DS${},$DS${}-1)".format(r, r, r)
    ws['FU{}'.format(s)] = "=STDEV(FU2:FU{})".format(q)
    ws['FU{}'.format(t)] = "=MAX(FU2:FU{})".format(q)
    ws['FU{}'.format(u)] = "=MIN(FU2:FU{})".format(q)
    # rata" KIM ke KIM tambahan dan mapel KIM awal
    ws['FV{}'.format(r)] = "=IF($FE${}=0,$DT${},$DT${}-1)".format(r, r, r)
    ws['FV{}'.format(s)] = "=STDEV(FV2:FV{})".format(q)
    ws['FV{}'.format(t)] = "=MAX(FV2:FV{})".format(q)
    ws['FV{}'.format(u)] = "=MIN(FV2:FV{})".format(q)
    # rata" BIO ke BIO tambahan dan mapel BIO awal
    ws['FW{}'.format(r)] = "=IF($FF${}=0,$DU${},$DU${}-1)".format(r, r, r)
    ws['FW{}'.format(s)] = "=STDEV(FW2:FW{})".format(q)
    ws['FW{}'.format(t)] = "=MAX(FW2:FW{})".format(q)
    ws['FW{}'.format(u)] = "=MIN(FW2:FW{})".format(q)
    # jml MAPEL
    ws['FX{}'.format(r)] = "=ROUND(AVERAGE(FX2:FX{}),2)".format(q)
    ws['FX{}'.format(t)] = "=MAX(FX2:FX{})".format(q)
    ws['FX{}'.format(u)] = "=MIN(FX2:FX{})".format(q)
    # MAX Z SCORE
    ws['FY{}'.format(r)] = "=MAX(FY2:FY{})".format(q)
    ws['FZ{}'.format(r)] = "=MAX(FZ2:FZ{})".format(q)
    ws['GA{}'.format(r)] = "=MAX(GA2:GA{})".format(q)
    ws['GB{}'.format(r)] = "=MAX(GB2:GB{})".format(q)
    ws['GC{}'.format(r)] = "=MAX(GC2:GC{})".format(q)
    ws['GD{}'.format(r)] = "=MAX(GD2:GD{})".format(q)
    ws['GE{}'.format(r)] = "=MAX(GE2:GE{})".format(q)
    ws['GF{}'.format(r)] = "=MAX(GF2:GF{})".format(q)
    ws['GG{}'.format(r)] = "=MAX(GG2:GG{})".format(q)
    ws['GH{}'.format(r)] = "=MAX(GH2:GH{})".format(q)
    ws['GI{}'.format(r)] = "=MAX(GI2:GI{})".format(q)
    # NILAI STANDAR MTW
    ws['GJ{}'.format(r)] = "=MAX(GJ2:GJ{})".format(q)
    ws['GJ{}'.format(s)] = "=MIN(GJ2:GJ{})".format(q)
    ws['GJ{}'.format(t)] = "=ROUND(AVERAGE(GJ2:GJ{}),2)".format(q)
    # NILAI STANDAR MTP
    ws['GK{}'.format(r)] = "=MAX(GK2:GK{})".format(q)
    ws['GK{}'.format(s)] = "=MIN(GK2:GK{})".format(q)
    ws['GK{}'.format(t)] = "=ROUND(AVERAGE(GK2:GK{}),2)".format(q)
    # NILAI STANDAR IND
    ws['GL{}'.format(r)] = "=MAX(GL2:GL{})".format(q)
    ws['GL{}'.format(s)] = "=MIN(GL2:GL{})".format(q)
    ws['GL{}'.format(t)] = "=ROUND(AVERAGE(GL2:GL{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['GM{}'.format(r)] = "=MAX(GM2:GM{})".format(q)
    ws['GM{}'.format(s)] = "=MIN(GM2:GM{})".format(q)
    ws['GM{}'.format(t)] = "=ROUND(AVERAGE(GM2:GM{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['GN{}'.format(r)] = "=MAX(GN2:GN{})".format(q)
    ws['GN{}'.format(s)] = "=MIN(GN2:GN{})".format(q)
    ws['GN{}'.format(t)] = "=ROUND(AVERAGE(GN2:GN{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['GO{}'.format(r)] = "=MAX(GO2:GO{})".format(q)
    ws['GO{}'.format(s)] = "=MIN(GO2:GO{})".format(q)
    ws['GO{}'.format(t)] = "=ROUND(AVERAGE(GO2:GO{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['GP{}'.format(r)] = "=MAX(GP2:GP{})".format(q)
    ws['GP{}'.format(s)] = "=MIN(GP2:GP{})".format(q)
    ws['GP{}'.format(t)] = "=ROUND(AVERAGE(GP2:GP{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['GQ{}'.format(r)] = "=MAX(GQ2:GQ{})".format(q)
    ws['GQ{}'.format(s)] = "=MIN(GQ2:GQ{})".format(q)
    ws['GQ{}'.format(t)] = "=ROUND(AVERAGE(GQ2:GQ{}),2)".format(q)
    # NILAI STANDAR FIS
    ws['GR{}'.format(r)] = "=MAX(GR2:GR{})".format(q)
    ws['GR{}'.format(s)] = "=MIN(GR2:GR{})".format(q)
    ws['GR{}'.format(t)] = "=ROUND(AVERAGE(GR2:GR{}),2)".format(q)
    # NILAI STANDAR KIM
    ws['GS{}'.format(r)] = "=MAX(GS2:GS{})".format(q)
    ws['GS{}'.format(s)] = "=MIN(GS2:GS{})".format(q)
    ws['GS{}'.format(t)] = "=ROUND(AVERAGE(GS2:GS{}),2)".format(q)
    # NILAI STANDAR BIO
    ws['GT{}'.format(r)] = "=MAX(GT2:GT{})".format(q)
    ws['GT{}'.format(s)] = "=MIN(GT2:GT{})".format(q)
    ws['GT{}'.format(t)] = "=ROUND(AVERAGE(GT2:GT{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['GU{}'.format(r)] = "=MAX(GU2:GU{})".format(q)
    ws['GU{}'.format(s)] = "=MIN(GU2:GU{})".format(q)
    ws['GU{}'.format(t)] = "=ROUND(AVERAGE(GU2:GU{}),2)".format(q)

    # TAMBAHAN
    ws['GX{}'.format(r)] = "=SUM(GX2:GX{})".format(q)
    ws['GY{}'.format(r)] = "=SUM(GY2:GY{})".format(q)
    ws['GZ{}'.format(r)] = "=SUM(GZ2:GZ{})".format(q)
    ws['HA{}'.format(r)] = "=SUM(HA2:HA{})".format(q)
    ws['HB{}'.format(r)] = "=SUM(HB2:HB{})".format(q)
    ws['HC{}'.format(r)] = "=SUM(HC2:HC{})".format(q)
    ws['HD{}'.format(r)] = "=SUM(HD2:HD{})".format(q)
    ws['HE{}'.format(r)] = "=SUM(HE2:HE{})".format(q)
    ws['HF{}'.format(r)] = "=SUM(HF2:HF{})".format(q)
    ws['HG{}'.format(r)] = "=SUM(HG2:HG{})".format(q)
    ws['HH{}'.format(r)] = "=SUM(HH2:HH{})".format(q)

    # iterasi 4 rata-rata - 4
    # rata" MTW ke MTW tambahan dan mapel MTW awal
    ws['HO{}'.format(r)] = "=IF($GX${}=0,$FM${},$FM${}-1)".format(r, r, r)
    ws['HO{}'.format(s)] = "=STDEV(HO2:HO{})".format(q)
    ws['HO{}'.format(t)] = "=MAX(HO2:HO{})".format(q)
    ws['HO{}'.format(u)] = "=MIN(HO2:HO{})".format(q)
    # rata" MTP ke MTP tambahan dan mapel MTP awal
    ws['HP{}'.format(r)] = "=IF($GY${}=0,$FN${},$FN${}-1)".format(r, r, r)
    ws['HP{}'.format(s)] = "=STDEV(HP2:HP{})".format(q)
    ws['HP{}'.format(t)] = "=MAX(HP2:HP{})".format(q)
    ws['HP{}'.format(u)] = "=MIN(HP2:HP{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['HQ{}'.format(r)] = "=IF($GZ${}=0,$FO${},$FO${}-1)".format(r, r, r)
    ws['HQ{}'.format(s)] = "=STDEV(HQ2:HQ{})".format(q)
    ws['HQ{}'.format(t)] = "=MAX(HQ2:HQ{})".format(q)
    ws['HQ{}'.format(u)] = "=MIN(HQ2:HQ{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['HR{}'.format(r)] = "=IF($HA${}=0,$FP${},$FP${}-1)".format(r, r, r)
    ws['HR{}'.format(s)] = "=STDEV(HR2:HR{})".format(q)
    ws['HR{}'.format(t)] = "=MAX(HR2:HR{})".format(q)
    ws['HR{}'.format(u)] = "=MIN(HR2:HR{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['HS{}'.format(r)] = "=IF($HB${}=0,$FQ${},$FQ${}-1)".format(r, r, r)
    ws['HS{}'.format(s)] = "=STDEV(HS2:HS{})".format(q)
    ws['HS{}'.format(t)] = "=MAX(HS2:HS{})".format(q)
    ws['HS{}'.format(u)] = "=MIN(HS2:HS{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['HT{}'.format(r)] = "=IF($HC${}=0,$FR${},$FR${}-1)".format(r, r, r)
    ws['HT{}'.format(s)] = "=STDEV(HT2:HT{})".format(q)
    ws['HT{}'.format(t)] = "=MAX(HT2:HT{})".format(q)
    ws['HT{}'.format(u)] = "=MIN(HT2:HT{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['HU{}'.format(r)] = "=IF($HD${}=0,$FS${},$FS${}-1)".format(r, r, r)
    ws['HU{}'.format(s)] = "=STDEV(HU2:HU{})".format(q)
    ws['HU{}'.format(t)] = "=MAX(HU2:HU{})".format(q)
    ws['HU{}'.format(u)] = "=MIN(HU2:HU{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['HV{}'.format(r)] = "=IF($HE${}=0,$FT${},$FT${}-1)".format(r, r, r)
    ws['HV{}'.format(s)] = "=STDEV(HV2:HV{})".format(q)
    ws['HV{}'.format(t)] = "=MAX(HV2:HV{})".format(q)
    ws['HV{}'.format(u)] = "=MIN(HV2:HV{})".format(q)
    # rata" FIS ke FIS tambahan dan mapel FIS awal
    ws['HW{}'.format(r)] = "=IF($HF${}=0,$FU${},$FU${}-1)".format(r, r, r)
    ws['HW{}'.format(s)] = "=STDEV(HW2:HW{})".format(q)
    ws['HW{}'.format(t)] = "=MAX(HW2:HW{})".format(q)
    ws['HW{}'.format(u)] = "=MIN(HW2:HW{})".format(q)
    # rata" KIM ke KIM tambahan dan mapel KIM awal
    ws['HX{}'.format(r)] = "=IF($HG${}=0,$FV${},$FV${}-1)".format(r, r, r)
    ws['HX{}'.format(s)] = "=STDEV(HX2:HX{})".format(q)
    ws['HX{}'.format(t)] = "=MAX(HX2:HX{})".format(q)
    ws['HX{}'.format(u)] = "=MIN(HX2:HX{})".format(q)
    # rata" BIO ke BIO tambahan dan mapel BIO awal
    ws['HY{}'.format(r)] = "=IF($HH${}=0,$FW${},$FW${}-1)".format(r, r, r)
    ws['HY{}'.format(s)] = "=STDEV(HY2:HY{})".format(q)
    ws['HY{}'.format(t)] = "=MAX(HY2:HY{})".format(q)
    ws['HY{}'.format(u)] = "=MIN(HY2:HY{})".format(q)
    # jml MAPEL
    ws['HZ{}'.format(r)] = "=ROUND(AVERAGE(HZ2:HZ{}),2)".format(q)
    ws['HZ{}'.format(t)] = "=MAX(HZ2:HZ{})".format(q)
    ws['HZ{}'.format(u)] = "=MIN(HZ2:HZ{})".format(q)
    # MAX Z SCORE
    ws['IA{}'.format(r)] = "=MAX(IA2:IA{})".format(q)
    ws['IB{}'.format(r)] = "=MAX(IB2:IB{})".format(q)
    ws['IC{}'.format(r)] = "=MAX(IC2:IC{})".format(q)
    ws['ID{}'.format(r)] = "=MAX(ID2:ID{})".format(q)
    ws['IE{}'.format(r)] = "=MAX(IE2:IE{})".format(q)
    ws['IF{}'.format(r)] = "=MAX(IF2:IF{})".format(q)
    ws['IG{}'.format(r)] = "=MAX(IG2:IG{})".format(q)
    ws['IH{}'.format(r)] = "=MAX(IH2:IH{})".format(q)
    ws['II{}'.format(r)] = "=MAX(II2:II{})".format(q)
    ws['IJ{}'.format(r)] = "=MAX(IJ2:IJ{})".format(q)
    ws['IK{}'.format(r)] = "=MAX(IK2:IK{})".format(q)
    # NILAI STANDAR MTA
    ws['IL{}'.format(r)] = "=MAX(IL2:IL{})".format(q)
    ws['IL{}'.format(s)] = "=MIN(IL2:IL{})".format(q)
    ws['IL{}'.format(t)] = "=ROUND(AVERAGE(IL2:IL{}),2)".format(q)
    # NILAI STANDAR MTP
    ws['IM{}'.format(r)] = "=MAX(IM2:IM{})".format(q)
    ws['IM{}'.format(s)] = "=MIN(IM2:IM{})".format(q)
    ws['IM{}'.format(t)] = "=ROUND(AVERAGE(IM2:IM{}),2)".format(q)
    # NILAI STANDAR IND
    ws['IN{}'.format(r)] = "=MAX(IN2:IN{})".format(q)
    ws['IN{}'.format(s)] = "=MIN(IN2:IN{})".format(q)
    ws['IN{}'.format(t)] = "=ROUND(AVERAGE(IN2:IN{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['IO{}'.format(r)] = "=MAX(IO2:IO{})".format(q)
    ws['IO{}'.format(s)] = "=MIN(IO2:IO{})".format(q)
    ws['IO{}'.format(t)] = "=ROUND(AVERAGE(IO2:IO{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['IP{}'.format(r)] = "=MAX(IP2:IP{})".format(q)
    ws['IP{}'.format(s)] = "=MIN(IP2:IP{})".format(q)
    ws['IP{}'.format(t)] = "=ROUND(AVERAGE(IP2:IP{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['IQ{}'.format(r)] = "=MAX(IQ2:IQ{})".format(q)
    ws['IQ{}'.format(s)] = "=MIN(IQ2:IQ{})".format(q)
    ws['IQ{}'.format(t)] = "=ROUND(AVERAGE(IQ2:IQ{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['IR{}'.format(r)] = "=MAX(IR2:IR{})".format(q)
    ws['IR{}'.format(s)] = "=MIN(IR2:IR{})".format(q)
    ws['IR{}'.format(t)] = "=ROUND(AVERAGE(IR2:IR{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['IS{}'.format(r)] = "=MAX(IS2:IS{})".format(q)
    ws['IS{}'.format(s)] = "=MIN(IS2:IS{})".format(q)
    ws['IS{}'.format(t)] = "=ROUND(AVERAGE(IS2:IS{}),2)".format(q)
    # NILAI STANDAR FIS
    ws['IT{}'.format(r)] = "=MAX(IT2:IT{})".format(q)
    ws['IT{}'.format(s)] = "=MIN(IT2:IT{})".format(q)
    ws['IT{}'.format(t)] = "=ROUND(AVERAGE(IT2:IT{}),2)".format(q)
    # NILAI STANDAR KIM
    ws['IU{}'.format(r)] = "=MAX(IU2:IU{})".format(q)
    ws['IU{}'.format(s)] = "=MIN(IU2:IU{})".format(q)
    ws['IU{}'.format(t)] = "=ROUND(AVERAGE(IU2:IU{}),2)".format(q)
    # NILAI STANDAR BIO
    ws['IV{}'.format(r)] = "=MAX(IV2:IV{})".format(q)
    ws['IV{}'.format(s)] = "=MIN(IV2:IV{})".format(q)
    ws['IV{}'.format(t)] = "=ROUND(AVERAGE(IV2:IV{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['IW{}'.format(r)] = "=MAX(IW2:IW{})".format(q)
    ws['IW{}'.format(s)] = "=MIN(IW2:IW{})".format(q)
    ws['IW{}'.format(t)] = "=ROUND(AVERAGE(IW2:IW{}),2)".format(q)

    # TAMBAHAN
    ws['IZ{}'.format(r)] = "=SUM(IZ2:IZ{})".format(q)
    ws['JA{}'.format(r)] = "=SUM(JA2:JA{})".format(q)
    ws['JB{}'.format(r)] = "=SUM(JB2:JB{})".format(q)
    ws['JC{}'.format(r)] = "=SUM(JC2:JC{})".format(q)
    ws['JD{}'.format(r)] = "=SUM(JD2:JD{})".format(q)
    ws['JE{}'.format(r)] = "=SUM(JE2:JE{})".format(q)
    ws['JF{}'.format(r)] = "=SUM(JF2:JF{})".format(q)
    ws['JG{}'.format(r)] = "=SUM(JG2:JG{})".format(q)
    ws['JH{}'.format(r)] = "=SUM(JH2:JH{})".format(q)
    ws['JI{}'.format(r)] = "=SUM(JI2:JI{})".format(q)
    ws['JJ{}'.format(r)] = "=SUM(JJ2:JJ{})".format(q)

    # iterasi 5 rata-rata - 5
    # rata" MTW ke MTW tambahan dan mapel MTW awal
    ws['JQ{}'.format(r)] = "=IF($IZ${}=0,$HO${},$HO${}-1)".format(r, r, r)
    ws['JQ{}'.format(s)] = "=STDEV(JQ2:JQ{})".format(q)
    ws['JQ{}'.format(t)] = "=MAX(JQ2:JQ{})".format(q)
    ws['JQ{}'.format(u)] = "=MIN(JQ2:JQ{})".format(q)
    # rata" MTP ke MTP tambahan dan mapel MTP awal
    ws['JR{}'.format(r)] = "=IF($JA${}=0,$HP${},$HP${}-1)".format(r, r, r)
    ws['JR{}'.format(s)] = "=STDEV(JR2:JR{})".format(q)
    ws['JR{}'.format(t)] = "=MAX(JR2:JR{})".format(q)
    ws['JR{}'.format(u)] = "=MIN(JR2:JR{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['JS{}'.format(r)] = "=IF($JB${}=0,$HQ${},$HQ${}-1)".format(r, r, r)
    ws['JS{}'.format(s)] = "=STDEV(JS2:JS{})".format(q)
    ws['JS{}'.format(t)] = "=MAX(JS2:JS{})".format(q)
    ws['JS{}'.format(u)] = "=MIN(JS2:JS{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['JT{}'.format(r)] = "=IF($JC${}=0,$HR${},$HR${}-1)".format(r, r, r)
    ws['JT{}'.format(s)] = "=STDEV(JT2:JT{})".format(q)
    ws['JT{}'.format(t)] = "=MAX(JT2:JT{})".format(q)
    ws['JT{}'.format(u)] = "=MIN(JT2:JT{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['JU{}'.format(r)] = "=IF($JD${}=0,$HS${},$HS${}-1)".format(r, r, r)
    ws['JU{}'.format(s)] = "=STDEV(JU2:JU{})".format(q)
    ws['JU{}'.format(t)] = "=MAX(JU2:JU{})".format(q)
    ws['JU{}'.format(u)] = "=MIN(JU2:JU{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['JV{}'.format(r)] = "=IF($JE${}=0,$HT${},$HT${}-1)".format(r, r, r)
    ws['JV{}'.format(s)] = "=STDEV(JV2:JV{})".format(q)
    ws['JV{}'.format(t)] = "=MAX(JV2:JV{})".format(q)
    ws['JV{}'.format(u)] = "=MIN(JV2:JV{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['JW{}'.format(r)] = "=IF($JF${}=0,$HU${},$HU${}-1)".format(r, r, r)
    ws['JW{}'.format(s)] = "=STDEV(JW2:JW{})".format(q)
    ws['JW{}'.format(t)] = "=MAX(JW2:JW{})".format(q)
    ws['JW{}'.format(u)] = "=MIN(JW2:JW{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['JX{}'.format(r)] = "=IF($JG${}=0,$HV${},$HV${}-1)".format(r, r, r)
    ws['JX{}'.format(s)] = "=STDEV(JX2:JX{})".format(q)
    ws['JX{}'.format(t)] = "=MAX(JX2:JX{})".format(q)
    ws['JX{}'.format(u)] = "=MIN(JX2:JX{})".format(q)
    # rata" FIS ke FIS tambahan dan mapel FIS awal
    ws['JY{}'.format(r)] = "=IF($JH${}=0,$HW${},$HW${}-1)".format(r, r, r)
    ws['JY{}'.format(s)] = "=STDEV(JY2:JY{})".format(q)
    ws['JY{}'.format(t)] = "=MAX(JY2:JY{})".format(q)
    ws['JY{}'.format(u)] = "=MIN(JY2:JY{})".format(q)
    # rata" KIM ke KIM tambahan dan mapel KIM awal
    ws['JZ{}'.format(r)] = "=IF($JI${}=0,$HX${},$HX${}-1)".format(r, r, r)
    ws['JZ{}'.format(s)] = "=STDEV(JZ2:JZ{})".format(q)
    ws['JZ{}'.format(t)] = "=MAX(JZ2:JZ{})".format(q)
    ws['JZ{}'.format(u)] = "=MIN(JZ2:JZ{})".format(q)
    # rata" BIO ke BIO tambahan dan mapel BIO awal
    ws['KA{}'.format(r)] = "=IF($JJ${}=0,$HY${},$HY${}-1)".format(r, r, r)
    ws['KA{}'.format(s)] = "=STDEV(KA2:KA{})".format(q)
    ws['KA{}'.format(t)] = "=MAX(KA2:KA{})".format(q)
    ws['KA{}'.format(u)] = "=MIN(KA2:KA{})".format(q)
    # jml MAPEL
    ws['KB{}'.format(r)] = "=ROUND(AVERAGE(KB2:KB{}),2)".format(q)
    ws['KB{}'.format(t)] = "=MAX(KB2:KB{})".format(q)
    ws['KB{}'.format(u)] = "=MIN(KB2:KB{})".format(q)
    # MAX Z SCORE
    ws['KC{}'.format(r)] = "=MAX(KC2:KC{})".format(q)
    ws['KD{}'.format(r)] = "=MAX(KD2:KD{})".format(q)
    ws['KE{}'.format(r)] = "=MAX(KE2:KE{})".format(q)
    ws['KF{}'.format(r)] = "=MAX(KF2:KF{})".format(q)
    ws['KG{}'.format(r)] = "=MAX(KG2:KG{})".format(q)
    ws['KH{}'.format(r)] = "=MAX(KH2:KH{})".format(q)
    ws['KI{}'.format(r)] = "=MAX(KI2:KI{})".format(q)
    ws['KJ{}'.format(r)] = "=MAX(KJ2:KJ{})".format(q)
    ws['KK{}'.format(r)] = "=MAX(KK2:KK{})".format(q)
    ws['KL{}'.format(r)] = "=MAX(KL2:KL{})".format(q)
    ws['KM{}'.format(r)] = "=MAX(KM2:KM{})".format(q)
    # NILAI STANDAR MTA
    ws['KN{}'.format(r)] = "=MAX(KN2:KN{})".format(q)
    ws['KN{}'.format(s)] = "=MIN(KN2:KN{})".format(q)
    ws['KN{}'.format(t)] = "=ROUND(AVERAGE(KN2:KN{}),2)".format(q)
    # NILAI STANDAR MTP
    ws['KO{}'.format(r)] = "=MAX(KO2:KO{})".format(q)
    ws['KO{}'.format(s)] = "=MIN(KO2:KO{})".format(q)
    ws['KO{}'.format(t)] = "=ROUND(AVERAGE(KO2:KO{}),2)".format(q)
    # NILAI STANDAR IND
    ws['KP{}'.format(r)] = "=MAX(KP2:KP{})".format(q)
    ws['KP{}'.format(s)] = "=MIN(KP2:KP{})".format(q)
    ws['KP{}'.format(t)] = "=ROUND(AVERAGE(KP2:KP{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['KQ{}'.format(r)] = "=MAX(KQ2:KQ{})".format(q)
    ws['KQ{}'.format(s)] = "=MIN(KQ2:KQ{})".format(q)
    ws['KQ{}'.format(t)] = "=ROUND(AVERAGE(KQ2:KQ{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['KR{}'.format(r)] = "=MAX(KR2:KR{})".format(q)
    ws['KR{}'.format(s)] = "=MIN(KR2:KR{})".format(q)
    ws['KR{}'.format(t)] = "=ROUND(AVERAGE(KR2:KR{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['KS{}'.format(r)] = "=MAX(KS2:KS{})".format(q)
    ws['KS{}'.format(s)] = "=MIN(KS2:KS{})".format(q)
    ws['KS{}'.format(t)] = "=ROUND(AVERAGE(KS2:KS{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['KT{}'.format(r)] = "=MAX(KT2:KT{})".format(q)
    ws['KT{}'.format(s)] = "=MIN(KT2:KT{})".format(q)
    ws['KT{}'.format(t)] = "=ROUND(AVERAGE(KT2:KT{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['KU{}'.format(r)] = "=MAX(KU2:KU{})".format(q)
    ws['KU{}'.format(s)] = "=MIN(KU2:KU{})".format(q)
    ws['KU{}'.format(t)] = "=ROUND(AVERAGE(KU2:KU{}),2)".format(q)
    # NILAI STANDAR FIS
    ws['KV{}'.format(r)] = "=MAX(KV2:KV{})".format(q)
    ws['KV{}'.format(s)] = "=MIN(KV2:KV{})".format(q)
    ws['KV{}'.format(t)] = "=ROUND(AVERAGE(KV2:KV{}),2)".format(q)
    # NILAI STANDAR KIM
    ws['KW{}'.format(r)] = "=MAX(KW2:KW{})".format(q)
    ws['KW{}'.format(s)] = "=MIN(KW2:KW{})".format(q)
    ws['KW{}'.format(t)] = "=ROUND(AVERAGE(KW2:KW{}),2)".format(q)
    # NILAI STANDAR BIO
    ws['KX{}'.format(r)] = "=MAX(KX2:KX{})".format(q)
    ws['KX{}'.format(s)] = "=MIN(KX2:KX{})".format(q)
    ws['KX{}'.format(t)] = "=ROUND(AVERAGE(KX2:KX{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['KY{}'.format(r)] = "=MAX(KY2:KY{})".format(q)
    ws['KY{}'.format(s)] = "=MIN(KY2:KY{})".format(q)
    ws['KY{}'.format(t)] = "=ROUND(AVERAGE(KY2:KY{}),2)".format(q)

    # TAMBAHAN
    ws['LB{}'.format(r)] = "=SUM(LB2:LB{})".format(q)
    ws['LC{}'.format(r)] = "=SUM(LC2:LC{})".format(q)
    ws['LD{}'.format(r)] = "=SUM(LD2:LD{})".format(q)
    ws['LE{}'.format(r)] = "=SUM(LE2:LE{})".format(q)
    ws['LF{}'.format(r)] = "=SUM(LF2:LF{})".format(q)
    ws['LG{}'.format(r)] = "=SUM(LG2:LG{})".format(q)
    ws['LH{}'.format(r)] = "=SUM(LH2:LH{})".format(q)
    ws['LI{}'.format(r)] = "=SUM(LI2:LI{})".format(q)
    ws['LJ{}'.format(r)] = "=SUM(LJ2:LJ{})".format(q)
    ws['LK{}'.format(r)] = "=SUM(LK2:LK{})".format(q)
    ws['LL{}'.format(r)] = "=SUM(LL2:LL{})".format(q)

    # Z Score [1]
    ws['B1'] = 'NAMA SISWA_A'
    ws['C1'] = 'NOMOR NF_A'
    ws['D1'] = 'KELAS_A'
    ws['E1'] = 'NAMA SEKOLAH_A'
    ws['F1'] = 'LOKASI_A'

    ws['G1'] = 'MAW_A'
    ws['H1'] = 'MAP_A'
    ws['I1'] = 'IND_A'
    ws['J1'] = 'ENG_A'
    ws['K1'] = 'SEJ_A'
    ws['L1'] = 'GEO_A'
    ws['M1'] = 'EKO_A'
    ws['N1'] = 'SOS_A'
    ws['O1'] = 'FIS_A'
    ws['P1'] = 'KIM_A'
    ws['Q1'] = 'BIO_A'
    ws['R1'] = 'JML_A'

    ws['S1'] = 'Z_MAW_A'
    ws['T1'] = 'Z_MAP_A'
    ws['U1'] = 'Z_IND_A'
    ws['V1'] = 'Z_ENG_A'
    ws['W1'] = 'Z_SEJ_A'
    ws['X1'] = 'Z_GEO_A'
    ws['Y1'] = 'Z_EKO_A'
    ws['Z1'] = 'Z_SOS_A'
    ws['AA1'] = 'Z_FIS_A'
    ws['AB1'] = 'Z_KIM_A'
    ws['AC1'] = 'Z_BIO_A'

    ws['AD1'] = 'S_MAW_A'
    ws['AE1'] = 'S_MAP_A'
    ws['AF1'] = 'S_IND_A'
    ws['AG1'] = 'S_ENG_A'
    ws['AH1'] = 'S_SEJ_A'
    ws['AI1'] = 'S_GEO_A'
    ws['AJ1'] = 'S_EKO_A'
    ws['AK1'] = 'S_SOS_A'
    ws['AL1'] = 'S_FIS_A'
    ws['AM1'] = 'S_KIM_A'
    ws['AN1'] = 'S_BIO_A'
    ws['AO1'] = 'S_JML_A'

    ws['AP1'] = 'RANK NAS._A'
    ws['AQ1'] = 'RANK LOK._A'

    ws['S1'].font = Font(bold=False, name='Calibri', size=11)
    ws['T1'].font = Font(bold=False, name='Calibri', size=11)
    ws['U1'].font = Font(bold=False, name='Calibri', size=11)
    ws['V1'].font = Font(bold=False, name='Calibri', size=11)
    ws['W1'].font = Font(bold=False, name='Calibri', size=11)
    ws['X1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['B1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['C1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['D1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['E1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['F1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['G1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['H1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['I1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['J1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['K1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['L1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['M1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['N1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['O1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['P1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Q1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['R1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['S1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['T1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['U1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['V1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['W1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['X1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Y1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Z1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AA1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AB1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AC1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AD1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AE1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AF1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AG1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AH1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AI1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AJ1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AK1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AL1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AM1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AN1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AO1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AP1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AQ1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')

    # tambahan
    ws['AR1'] = 'MAW_20_A'
    ws['AS1'] = 'MAP_20_A'
    ws['AT1'] = 'IND_20_A'
    ws['AU1'] = 'ENG_20_A'
    ws['AV1'] = 'SEJ_20_A'
    ws['AW1'] = 'GEO_20_A'
    ws['AX1'] = 'EKO_20_A'
    ws['AY1'] = 'SOS_20_A'
    ws['AZ1'] = 'FIS_20_A'
    ws['BA1'] = 'KIM_20_A'
    ws['BB1'] = 'BIO_20_A'

    ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BB1'].font = Font(bold=False, name='Calibri', size=11)

    ws['AR1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AS1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AT1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AU1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AV1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AW1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AX1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AY1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AZ1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['BA1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['BB1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')

    for row in range(2, q+1):
        ws['S{}'.format(
            row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
        ws['T{}'.format(
            row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
        ws['U{}'.format(
            row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
        ws['V{}'.format(
            row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
        ws['W{}'.format(
            row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
        ws['X{}'.format(
            row)] = '=IFERROR(ROUND(IF(L{}="","",(L{}-L${})/L${}),2),"")'.format(row, row, r, s)
        ws['Y{}'.format(
            row)] = '=IFERROR(ROUND(IF(M{}="","",(M{}-M${})/M${}),2),"")'.format(row, row, r, s)
        ws['Z{}'.format(
            row)] = '=IFERROR(ROUND(IF(N{}="","",(N{}-N${})/N${}),2),"")'.format(row, row, r, s)
        ws['AA{}'.format(
            row)] = '=IFERROR(ROUND(IF(O{}="","",(O{}-O${})/O${}),2),"")'.format(row, row, r, s)
        ws['AB{}'.format(
            row)] = '=IFERROR(ROUND(IF(P{}="","",(P{}-P${})/P${}),2),"")'.format(row, row, r, s)
        ws['AC{}'.format(
            row)] = '=IFERROR(ROUND(IF(Q{}="","",(Q{}-Q${})/Q${}),2),"")'.format(row, row, r, s)

        ws['AD{}'.format(
            row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*S{}/$S${}<20,20,70+30*S{}/$S${})),2),"")'.format(row, row, r, row, r)
        ws['AE{}'.format(
            row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*T{}/$T${}<20,20,70+30*T{}/$T${})),2),"")'.format(row, row, r, row, r)
        ws['AF{}'.format(
            row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*U{}/$U${}<20,20,70+30*U{}/$U${})),2),"")'.format(row, row, r, row, r)
        ws['AG{}'.format(
            row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*V{}/$V${}<20,20,70+30*V{}/$V${})),2),"")'.format(row, row, r, row, r)
        ws['AH{}'.format(
            row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*W{}/$W${}<20,20,70+30*W{}/$W${})),2),"")'.format(row, row, r, row, r)
        ws['AI{}'.format(
            row)] = '=IFERROR(ROUND(IF(L{}="","",IF(70+30*X{}/$X${}<20,20,70+30*X{}/$X${})),2),"")'.format(row, row, r, row, r)
        ws['AJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(M{}="","",IF(70+30*Y{}/$Y${}<20,20,70+30*Y{}/$Y${})),2),"")'.format(row, row, r, row, r)
        ws['AK{}'.format(
            row)] = '=IFERROR(ROUND(IF(N{}="","",IF(70+30*Z{}/$Z${}<20,20,70+30*Z{}/$Z${})),2),"")'.format(row, row, r, row, r)
        ws['AL{}'.format(
            row)] = '=IFERROR(ROUND(IF(O{}="","",IF(70+30*AA{}/$AA${}<20,20,70+30*AA{}/$AA${})),2),"")'.format(row, row, r, row, r)
        ws['AM{}'.format(
            row)] = '=IFERROR(ROUND(IF(P{}="","",IF(70+30*AB{}/$AB${}<20,20,70+30*AB{}/$AB${})),2),"")'.format(row, row, r, row, r)
        ws['AN{}'.format(
            row)] = '=IFERROR(ROUND(IF(Q{}="","",IF(70+30*AC{}/$AC${}<20,20,70+30*AC{}/$AC${})),2),"")'.format(row, row, r, row, r)

        ws['AO{}'.format(row)] = '=IF(SUM(AD{}:AN{})=0,"",SUM(AD{}:AN{}))'.format(
            row, row, row, row)
        ws['AP{}'.format(row)] = '=IF(AO{}="","",RANK(AO{},$AO$2:$AO${}))'.format(
            row, row, q)
        ws['AQ{}'.format(
            row)] = '=IF(AP{}="","",COUNTIFS($F$2:$F${},F{},$AP$2:$AP${},"<"&AP{})+1)'.format(row, q, row, q, row)

    # TAMBAHAN
        ws['AR{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,AD{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,AD{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,AD{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,AD{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,AD{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,AD{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AS{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,AE{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,AE{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,AE{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,AE{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,AE{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,AE{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AT{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,AF{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,AF{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,AF{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,AF{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,AF{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,AF{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AU{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,AG{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,AG{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,AG{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,AG{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,AG{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,AG{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AV{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,AH{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,AH{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,AH{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,AH{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,AH{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,AH{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AW{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,AI{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,AI{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,AI{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,AI{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,AI{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,AI{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AX{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,AJ{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,AJ{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,AJ{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,AJ{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,AJ{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,AJ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AY{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,AK{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,AK{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,AK{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,AK{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,AK{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,AK{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AZ{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,AL{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,AL{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,AL{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,AL{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,AL{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,AL{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BA{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,AM{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,AM{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,AM{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,AM{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,AM{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,AM{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BB{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,AN{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,AN{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,AN{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,AN{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,AN{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,AN{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [2]
    ws['BD1'] = 'NAMA SISWA_A'
    ws['BE1'] = 'NOMOR NF_A'
    ws['BF1'] = 'KELAS_A'
    ws['BG1'] = 'NAMA SEKOLAH_A'
    ws['BH1'] = 'LOKASI_A'

    ws['BI1'] = 'MAW_A'
    ws['BJ1'] = 'MAP_A'
    ws['BK1'] = 'IND_A'
    ws['BL1'] = 'ENG_A'
    ws['BM1'] = 'SEJ_A'
    ws['BN1'] = 'GEO_A'
    ws['BO1'] = 'EKO_A'
    ws['BP1'] = 'SOS_A'
    ws['BQ1'] = 'FIS_A'
    ws['BR1'] = 'KIM_A'
    ws['BS1'] = 'BIO_A'
    ws['BT1'] = 'JML_A'

    ws['BU1'] = 'Z_MAW_A'
    ws['BV1'] = 'Z_MAP_A'
    ws['BW1'] = 'Z_IND_A'
    ws['BX1'] = 'Z_ENG_A'
    ws['BY1'] = 'Z_SEJ_A'
    ws['BZ1'] = 'Z_GEO_A'
    ws['CA1'] = 'Z_EKO_A'
    ws['CB1'] = 'Z_SOS_A'
    ws['CC1'] = 'Z_FIS_A'
    ws['CD1'] = 'Z_KIM_A'
    ws['CE1'] = 'Z_BIO_A'

    ws['CF1'] = 'S_MAW_A'
    ws['CG1'] = 'S_MAP_A'
    ws['CH1'] = 'S_IND_A'
    ws['CI1'] = 'S_ENG_A'
    ws['CJ1'] = 'S_SEJ_A'
    ws['CK1'] = 'S_GEO_A'
    ws['CL1'] = 'S_EKO_A'
    ws['CM1'] = 'S_SOS_A'
    ws['CN1'] = 'S_FIS_A'
    ws['CO1'] = 'S_KIM_A'
    ws['CP1'] = 'S_BIO_A'
    ws['CQ1'] = 'S_JML_A'

    ws['CR1'] = 'RANK NAS._A'
    ws['CS1'] = 'RANK LOK._A'

    ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CS1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['BD1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BE1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BF1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BG1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BH1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BI1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BJ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BK1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BL1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BM1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BN1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BO1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BP1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BQ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BR1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BS1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BT1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BU1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BV1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BW1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BX1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BY1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BZ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CA1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CB1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CC1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CD1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CE1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CF1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CG1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CH1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CI1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CJ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CK1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CL1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CM1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CN1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CO1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CP1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CQ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CR1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CS1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')

    # tambahan
    ws['CT1'] = 'MAW_20_A'
    ws['CU1'] = 'MAP_20_A'
    ws['CV1'] = 'IND_20_A'
    ws['CW1'] = 'ENG_20_A'
    ws['CX1'] = 'SEJ_20_A'
    ws['CY1'] = 'GEO_20_A'
    ws['CZ1'] = 'EKO_20_A'
    ws['DA1'] = 'SOS_20_A'
    ws['DB1'] = 'FIS_20_A'
    ws['DC1'] = 'KIM_20_A'
    ws['DD1'] = 'BIO_20_A'

    ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DD1'].font = Font(bold=False, name='Calibri', size=11)

    ws['CT1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CU1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CV1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CW1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CX1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CY1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['CZ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['DA1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['DB1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['DC1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['DD1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')

    for row in range(2, q+1):
        ws['BD{}'.format(row)] = '=B{}'.format(row)
        ws['BE{}'.format(row)] = '=C{}'.format(row, row)
        ws['BF{}'.format(row)] = '=D{}'.format(row, row)
        ws['BG{}'.format(row)] = '=E{}'.format(row, row)
        ws['BH{}'.format(row)] = '=F{}'.format(row, row)

        ws['BI{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
        ws['BJ{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
        ws['BK{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
        ws['BL{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
        ws['BM{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
        ws['BN{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
        ws['BO{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
        ws['BP{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
        ws['BQ{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
        ws['BR{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
        ws['BS{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
        ws['BT{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
        ws['BU{}'.format(
            row)] = '=IFERROR(ROUND(IF(BI{}="","",(BI{}-BI${})/BI${}),2),"")'.format(row, row, r, s)
        ws['BV{}'.format(
            row)] = '=IFERROR(ROUND(IF(BJ{}="","",(BJ{}-BJ${})/BJ${}),2),"")'.format(row, row, r, s)
        ws['BW{}'.format(
            row)] = '=IFERROR(ROUND(IF(BK{}="","",(BK{}-BK${})/BK${}),2),"")'.format(row, row, r, s)
        ws['BX{}'.format(
            row)] = '=IFERROR(ROUND(IF(BL{}="","",(BL{}-BL${})/BL${}),2),"")'.format(row, row, r, s)
        ws['BY{}'.format(
            row)] = '=IFERROR(ROUND(IF(BM{}="","",(BM{}-BM${})/BM${}),2),"")'.format(row, row, r, s)
        ws['BZ{}'.format(
            row)] = '=IFERROR(ROUND(IF(BN{}="","",(BN{}-BN${})/BN${}),2),"")'.format(row, row, r, s)
        ws['CA{}'.format(
            row)] = '=IFERROR(ROUND(IF(BO{}="","",(BO{}-BO${})/BO${}),2),"")'.format(row, row, r, s)
        ws['CB{}'.format(
            row)] = '=IFERROR(ROUND(IF(BP{}="","",(BP{}-BP${})/BP${}),2),"")'.format(row, row, r, s)
        ws['CC{}'.format(
            row)] = '=IFERROR(ROUND(IF(BQ{}="","",(BQ{}-BQ${})/BQ${}),2),"")'.format(row, row, r, s)
        ws['CD{}'.format(
            row)] = '=IFERROR(ROUND(IF(BR{}="","",(BR{}-BR${})/BR${}),2),"")'.format(row, row, r, s)
        ws['CE{}'.format(
            row)] = '=IFERROR(ROUND(IF(BS{}="","",(BS{}-BS${})/BS${}),2),"")'.format(row, row, r, s)

        ws['CF{}'.format(
            row)] = '=IFERROR(ROUND(IF(BI{}="","",IF(70+30*BU{}/$BU${}<20,20,70+30*BU{}/$BU${})),2),"")'.format(row, row, r, row, r)
        ws['CG{}'.format(
            row)] = '=IFERROR(ROUND(IF(BJ{}="","",IF(70+30*BV{}/$BV${}<20,20,70+30*BV{}/$BV${})),2),"")'.format(row, row, r, row, r)
        ws['CH{}'.format(
            row)] = '=IFERROR(ROUND(IF(BK{}="","",IF(70+30*BW{}/$BW${}<20,20,70+30*BW{}/$BW${})),2),"")'.format(row, row, r, row, r)
        ws['CI{}'.format(
            row)] = '=IFERROR(ROUND(IF(BL{}="","",IF(70+30*BX{}/$BX${}<20,20,70+30*BX{}/$BX${})),2),"")'.format(row, row, r, row, r)
        ws['CJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(BM{}="","",IF(70+30*BY{}/$BY${}<20,20,70+30*BY{}/$BY${})),2),"")'.format(row, row, r, row, r)
        ws['CK{}'.format(
            row)] = '=IFERROR(ROUND(IF(BN{}="","",IF(70+30*BZ{}/$BZ${}<20,20,70+30*BZ{}/$BZ${})),2),"")'.format(row, row, r, row, r)
        ws['CL{}'.format(
            row)] = '=IFERROR(ROUND(IF(BO{}="","",IF(70+30*CA{}/$CA${}<20,20,70+30*CA{}/$CA${})),2),"")'.format(row, row, r, row, r)
        ws['CM{}'.format(
            row)] = '=IFERROR(ROUND(IF(BP{}="","",IF(70+30*CB{}/$CB${}<20,20,70+30*CB{}/$CB${})),2),"")'.format(row, row, r, row, r)
        ws['CN{}'.format(
            row)] = '=IFERROR(ROUND(IF(BQ{}="","",IF(70+30*CC{}/$CC${}<20,20,70+30*CC{}/$CC${})),2),"")'.format(row, row, r, row, r)
        ws['CO{}'.format(
            row)] = '=IFERROR(ROUND(IF(BR{}="","",IF(70+30*CD{}/$CD${}<20,20,70+30*CD{}/$CD${})),2),"")'.format(row, row, r, row, r)
        ws['CP{}'.format(
            row)] = '=IFERROR(ROUND(IF(BS{}="","",IF(70+30*CE{}/$CE${}<20,20,70+30*CE{}/$CE${})),2),"")'.format(row, row, r, row, r)

        ws['CQ{}'.format(row)] = '=IF(SUM(CF{}:CP{})=0,"",SUM(CF{}:CP{}))'.format(
            row, row, row, row)
        ws['CR{}'.format(row)] = '=IF(CQ{}="","",RANK(CQ{},$CQ$2:$CQ${}))'.format(
            row, row, q)
        ws['CS{}'.format(
            row)] = '=IF(CR{}="","",COUNTIFS($BH$2:$BH${},BH{},$CR$2:$CR${},"<"&CR{})+1)'.format(row, q, row, q, row)

    # TAMBAHAN
        ws['CT{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,CF{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,CF{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,CF{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,CF{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,CF{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,CF{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CU{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,CG{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,CG{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,CG{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,CG{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,CG{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,CG{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CV{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,CH{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,CH{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,CH{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,CH{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,CH{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,CH{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CW{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,CI{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,CI{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,CI{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,CI{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,CI{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,CI{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CX{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,CJ{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,CJ{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,CJ{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,CJ{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,CJ{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,CJ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CY{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,CK{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,CK{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,CK{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,CK{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,CK{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,CK{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CZ{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,CL{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,CL{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,CL{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,CL{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,CL{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,CL{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DA{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,CM{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,CM{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,CM{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,CM{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,CM{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,CM{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DB{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,CN{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,CN{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,CN{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,CN{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,CN{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,CN{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DC{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,CO{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,CO{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,CO{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,CO{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,CO{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,CO{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DD{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,CP{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,CP{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,CP{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,CP{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,CP{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,CP{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [3]
    ws['DF1'] = 'NAMA SISWA_A'
    ws['DG1'] = 'NOMOR NF_A'
    ws['DH1'] = 'KELAS_A'
    ws['DI1'] = 'NAMA SEKOLAH_A'
    ws['DJ1'] = 'LOKASI_A'

    ws['DK1'] = 'MAW_A'
    ws['DL1'] = 'MAP_A'
    ws['DM1'] = 'IND_A'
    ws['DN1'] = 'ENG_A'
    ws['DO1'] = 'SEJ_A'
    ws['DP1'] = 'GEO_A'
    ws['DQ1'] = 'EKO_A'
    ws['DR1'] = 'SOS_A'
    ws['DS1'] = 'FIS_A'
    ws['DT1'] = 'KIM_A'
    ws['DU1'] = 'BIO_A'
    ws['DV1'] = 'JML_A'

    ws['DW1'] = 'Z_MAW_A'
    ws['DX1'] = 'Z_MAP_A'
    ws['DY1'] = 'Z_IND_A'
    ws['DZ1'] = 'Z_ENG_A'
    ws['EA1'] = 'Z_SEJ_A'
    ws['EB1'] = 'Z_GEO_A'
    ws['EC1'] = 'Z_EKO_A'
    ws['ED1'] = 'Z_SOS_A'
    ws['EE1'] = 'Z_FIS_A'
    ws['EF1'] = 'Z_KIM_A'
    ws['EG1'] = 'Z_BIO_A'

    ws['EH1'] = 'S_MAW_A'
    ws['EI1'] = 'S_MAP_A'
    ws['EJ1'] = 'S_IND_A'
    ws['EK1'] = 'S_ENG_A'
    ws['EL1'] = 'S_SEJ_A'
    ws['EM1'] = 'S_GEO_A'
    ws['EN1'] = 'S_EKO_A'
    ws['EO1'] = 'S_SOS_A'
    ws['EP1'] = 'S_FIS_A'
    ws['EQ1'] = 'S_KIM_A'
    ws['ER1'] = 'S_BIO_A'
    ws['ES1'] = 'S_JML_A'

    ws['ET1'] = 'RANK NAS._A'
    ws['EU1'] = 'RANK LOK._A'

    ws['DW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EU1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['DF1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DG1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DH1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DI1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DJ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DK1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DL1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DM1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DN1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DO1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DP1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DQ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DR1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DS1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DT1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DU1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DV1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DW1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DX1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DY1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DZ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EA1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EB1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EC1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['ED1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EE1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EF1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EG1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EH1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EI1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EJ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EK1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EL1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EM1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EN1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EO1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EP1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EQ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['ER1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['ES1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['ET1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EU1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')

    # tambahan
    ws['EV1'] = 'MAW_20_A'
    ws['EW1'] = 'MAP_20_A'
    ws['EX1'] = 'IND_20_A'
    ws['EY1'] = 'ENG_20_A'
    ws['EZ1'] = 'SEJ_20_A'
    ws['FA1'] = 'GEO_20_A'
    ws['FB1'] = 'EKO_20_A'
    ws['FC1'] = 'SOS_20_A'
    ws['FD1'] = 'FIS_20_A'
    ws['FE1'] = 'KIM_20_A'
    ws['FF1'] = 'BIO_20_A'

    ws['EV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FF1'].font = Font(bold=False, name='Calibri', size=11)

    ws['EV1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EW1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EX1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EY1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['EZ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['FA1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['FB1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['FC1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['FD1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['FE1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['FF1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')

    for row in range(2, q+1):
        ws['DF{}'.format(row)] = '=B{}'.format(row)
        ws['DG{}'.format(row)] = '=C{}'.format(row, row)
        ws['DH{}'.format(row)] = '=D{}'.format(row, row)
        ws['DI{}'.format(row)] = '=E{}'.format(row, row)
        ws['DJ{}'.format(row)] = '=F{}'.format(row, row)

        ws['DK{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
        ws['DL{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
        ws['DM{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
        ws['DN{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
        ws['DO{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
        ws['DP{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
        ws['DQ{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
        ws['DR{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
        ws['DS{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
        ws['DT{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
        ws['DU{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
        ws['DV{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
        ws['DW{}'.format(
            row)] = '=IFERROR(ROUND(IF(DK{}="","",(DK{}-DK${})/DK${}),2),"")'.format(row, row, r, s)
        ws['DX{}'.format(
            row)] = '=IFERROR(ROUND(IF(DL{}="","",(DL{}-DL${})/DL${}),2),"")'.format(row, row, r, s)
        ws['DY{}'.format(
            row)] = '=IFERROR(ROUND(IF(DM{}="","",(DM{}-DM${})/DM${}),2),"")'.format(row, row, r, s)
        ws['DZ{}'.format(
            row)] = '=IFERROR(ROUND(IF(DN{}="","",(DN{}-DN${})/DN${}),2),"")'.format(row, row, r, s)
        ws['EA{}'.format(
            row)] = '=IFERROR(ROUND(IF(DO{}="","",(DO{}-DO${})/DO${}),2),"")'.format(row, row, r, s)
        ws['EB{}'.format(
            row)] = '=IFERROR(ROUND(IF(DP{}="","",(DP{}-DP${})/DP${}),2),"")'.format(row, row, r, s)
        ws['EC{}'.format(
            row)] = '=IFERROR(ROUND(IF(DQ{}="","",(DQ{}-DQ${})/DQ${}),2),"")'.format(row, row, r, s)
        ws['ED{}'.format(
            row)] = '=IFERROR(ROUND(IF(DR{}="","",(DR{}-DR${})/DR${}),2),"")'.format(row, row, r, s)
        ws['EE{}'.format(
            row)] = '=IFERROR(ROUND(IF(DS{}="","",(DS{}-DS${})/DS${}),2),"")'.format(row, row, r, s)
        ws['EF{}'.format(
            row)] = '=IFERROR(ROUND(IF(DT{}="","",(DT{}-DT${})/DT${}),2),"")'.format(row, row, r, s)
        ws['EG{}'.format(
            row)] = '=IFERROR(ROUND(IF(DU{}="","",(DU{}-DU${})/DU${}),2),"")'.format(row, row, r, s)

        ws['EH{}'.format(
            row)] = '=IFERROR(ROUND(IF(DK{}="","",IF(70+30*DW{}/$DW${}<20,20,70+30*DW{}/$DW${})),2),"")'.format(row, row, r, row, r)
        ws['EI{}'.format(
            row)] = '=IFERROR(ROUND(IF(DL{}="","",IF(70+30*DX{}/$DX${}<20,20,70+30*DX{}/$DX${})),2),"")'.format(row, row, r, row, r)
        ws['EJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(DM{}="","",IF(70+30*DY{}/$DY${}<20,20,70+30*DY{}/$DY${})),2),"")'.format(row, row, r, row, r)
        ws['EK{}'.format(
            row)] = '=IFERROR(ROUND(IF(DN{}="","",IF(70+30*DZ{}/$DZ${}<20,20,70+30*DZ{}/$DZ${})),2),"")'.format(row, row, r, row, r)
        ws['EL{}'.format(
            row)] = '=IFERROR(ROUND(IF(DO{}="","",IF(70+30*EA{}/$EA${}<20,20,70+30*EA{}/$EA${})),2),"")'.format(row, row, r, row, r)
        ws['EM{}'.format(
            row)] = '=IFERROR(ROUND(IF(DP{}="","",IF(70+30*EB{}/$EB${}<20,20,70+30*EB{}/$EB${})),2),"")'.format(row, row, r, row, r)
        ws['EN{}'.format(
            row)] = '=IFERROR(ROUND(IF(DQ{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
        ws['EO{}'.format(
            row)] = '=IFERROR(ROUND(IF(DR{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
        ws['EP{}'.format(
            row)] = '=IFERROR(ROUND(IF(DS{}="","",IF(70+30*EF{}/$EF${}<20,20,70+30*EF{}/$EF${})),2),"")'.format(row, row, r, row, r)
        ws['EQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(DT{}="","",IF(70+30*EG{}/$EG${}<20,20,70+30*EG{}/$EG${})),2),"")'.format(row, row, r, row, r)
        ws['ER{}'.format(
            row)] = '=IFERROR(ROUND(IF(DU{}="","",IF(70+30*EH{}/$EH${}<20,20,70+30*EH{}/$EH${})),2),"")'.format(row, row, r, row, r)

        ws['ES{}'.format(row)] = '=IF(SUM(EH{}:ER{})=0,"",SUM(EH{}:ER{}))'.format(
            row, row, row, row)
        ws['ET{}'.format(row)] = '=IF(ES{}="","",RANK(ES{},$ES$2:$ES${}))'.format(
            row, row, q)
        ws['EU{}'.format(
            row)] = '=IF(ET{}="","",COUNTIFS($DJ$2:$DJ${},DJ{},$ET$2:$ET${},"<"&ET{})+1)'.format(row, q, row, q, row)

    # TAMBAHAN
        ws['EV{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,EH{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,EH{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,EH{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,EH{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,EH{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,EH{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EW{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,EI{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,EI{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,EI{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,EI{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,EI{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,EI{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EX{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,EJ{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,EJ{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,EJ{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,EJ{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,EJ{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,EJ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EY{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,EK{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,EK{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,EK{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,EK{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,EK{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,EK{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EZ{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,EL{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,EL{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,EL{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,EL{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,EL{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,EL{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['FA{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,EM{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,EM{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,EM{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,EM{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,EM{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,EM{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['FB{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,EN{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,EN{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,EN{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,EN{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,EN{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,EN{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['FC{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,EO{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,EO{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,EO{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,EO{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,EO{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,EO{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['FD{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,EP{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,EP{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,EP{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,EP{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,EP{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,EP{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['FE{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,EQ{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,EQ{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,EQ{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,EQ{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,EQ{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,EQ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['FF{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,ER{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,ER{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,ER{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,ER{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,ER{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,ER{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [4]
    ws['FH1'] = 'NAMA SISWA_A'
    ws['FI1'] = 'NOMOR NF_A'
    ws['FJ1'] = 'KELAS_A'
    ws['FK1'] = 'NAMA SEKOLAH_A'
    ws['FL1'] = 'LOKASI_A'

    ws['FM1'] = 'MAW_A'
    ws['FN1'] = 'MAP_A'
    ws['FO1'] = 'IND_A'
    ws['FP1'] = 'ENG_A'
    ws['FQ1'] = 'SEJ_A'
    ws['FR1'] = 'GEO_A'
    ws['FS1'] = 'EKO_A'
    ws['FT1'] = 'SOS_A'
    ws['FU1'] = 'FIS_A'
    ws['FV1'] = 'KIM_A'
    ws['FW1'] = 'BIO_A'
    ws['FX1'] = 'JML_A'

    ws['FY1'] = 'Z_MAW_A'
    ws['FZ1'] = 'Z_MAP_A'
    ws['GA1'] = 'Z_IND_A'
    ws['GB1'] = 'Z_ENG_A'
    ws['GC1'] = 'Z_SEJ_A'
    ws['GD1'] = 'Z_GEO_A'
    ws['GE1'] = 'Z_EKO_A'
    ws['GF1'] = 'Z_SOS_A'
    ws['GG1'] = 'Z_FIS_A'
    ws['GH1'] = 'Z_KIM_A'
    ws['GI1'] = 'Z_BIO_A'

    ws['GJ1'] = 'S_MAW_A'
    ws['GK1'] = 'S_MAP_A'
    ws['GL1'] = 'S_IND_A'
    ws['GM1'] = 'S_ENG_A'
    ws['GN1'] = 'S_SEJ_A'
    ws['GO1'] = 'S_GEO_A'
    ws['GP1'] = 'S_EKO_A'
    ws['GQ1'] = 'S_SOS_A'
    ws['GR1'] = 'S_FIS_A'
    ws['GS1'] = 'S_KIM_A'
    ws['GT1'] = 'S_BIO_A'
    ws['GU1'] = 'S_JML_A'

    ws['GV1'] = 'RANK NAS._A'
    ws['GW1'] = 'RANK LOK._A'

    ws['FY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GW1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['FH1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FI1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FJ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FK1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FL1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FM1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FN1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FO1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FP1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FQ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FR1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FS1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FT1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FU1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FV1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FW1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FX1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FY1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['FZ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GA1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GB1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GC1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GD1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GE1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GF1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GG1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GH1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GI1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GJ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GK1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GL1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GM1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GN1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GO1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GP1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GQ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GR1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GS1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GT1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GU1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GV1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GW1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

    # tambahan
    ws['GX1'] = 'MAW_20_A'
    ws['GY1'] = 'MAP_20_A'
    ws['GZ1'] = 'IND_20_A'
    ws['HA1'] = 'ENG_20_A'
    ws['HB1'] = 'SEJ_20_A'
    ws['HC1'] = 'GEO_20_A'
    ws['HD1'] = 'EKO_20_A'
    ws['HE1'] = 'SOS_20_A'
    ws['HF1'] = 'FIS_20_A'
    ws['HG1'] = 'KIM_20_A'
    ws['HH1'] = 'BIO_20_A'

    ws['GX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HH1'].font = Font(bold=False, name='Calibri', size=11)

    ws['GX1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GY1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['GZ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HA1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HB1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HC1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HD1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HE1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HF1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HG1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['HH1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

    for row in range(2, q+1):
        ws['FH{}'.format(row)] = '=B{}'.format(row)
        ws['FI{}'.format(row)] = '=C{}'.format(row, row)
        ws['FJ{}'.format(row)] = '=D{}'.format(row, row)
        ws['FK{}'.format(row)] = '=E{}'.format(row, row)
        ws['FL{}'.format(row)] = '=F{}'.format(row, row)

        ws['FM{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
        ws['FN{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
        ws['FO{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
        ws['FP{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
        ws['FQ{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
        ws['FR{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
        ws['FS{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
        ws['FT{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
        ws['FU{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
        ws['FV{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
        ws['FW{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
        ws['FX{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
        ws['FY{}'.format(
            row)] = '=IFERROR(ROUND(IF(FM{}="","",(FM{}-FM${})/FM${}),2),"")'.format(row, row, r, s)
        ws['FZ{}'.format(
            row)] = '=IFERROR(ROUND(IF(FN{}="","",(FN{}-FN${})/FN${}),2),"")'.format(row, row, r, s)
        ws['GA{}'.format(
            row)] = '=IFERROR(ROUND(IF(FO{}="","",(FO{}-FO${})/FO${}),2),"")'.format(row, row, r, s)
        ws['GB{}'.format(
            row)] = '=IFERROR(ROUND(IF(FP{}="","",(FP{}-FP${})/FP${}),2),"")'.format(row, row, r, s)
        ws['GC{}'.format(
            row)] = '=IFERROR(ROUND(IF(FQ{}="","",(FQ{}-FQ${})/FQ${}),2),"")'.format(row, row, r, s)
        ws['GD{}'.format(
            row)] = '=IFERROR(ROUND(IF(FR{}="","",(FR{}-FR${})/FR${}),2),"")'.format(row, row, r, s)
        ws['GE{}'.format(
            row)] = '=IFERROR(ROUND(IF(FS{}="","",(FS{}-FS${})/FS${}),2),"")'.format(row, row, r, s)
        ws['GF{}'.format(
            row)] = '=IFERROR(ROUND(IF(FT{}="","",(FT{}-FT${})/FT${}),2),"")'.format(row, row, r, s)
        ws['GG{}'.format(
            row)] = '=IFERROR(ROUND(IF(FU{}="","",(FU{}-FU${})/FU${}),2),"")'.format(row, row, r, s)
        ws['GH{}'.format(
            row)] = '=IFERROR(ROUND(IF(FV{}="","",(FV{}-FV${})/FV${}),2),"")'.format(row, row, r, s)
        ws['GI{}'.format(
            row)] = '=IFERROR(ROUND(IF(DU{}="","",(DU{}-DU${})/DU${}),2),"")'.format(row, row, r, s)

        ws['GJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(FM{}="","",IF(70+30*FY{}/$FY${}<20,20,70+30*FY{}/$FY${})),2),"")'.format(row, row, r, row, r)
        ws['GK{}'.format(
            row)] = '=IFERROR(ROUND(IF(FN{}="","",IF(70+30*FZ{}/$FZ${}<20,20,70+30*FZ{}/$FZ${})),2),"")'.format(row, row, r, row, r)
        ws['GL{}'.format(
            row)] = '=IFERROR(ROUND(IF(FO{}="","",IF(70+30*GA{}/$GA${}<20,20,70+30*GA{}/$GA${})),2),"")'.format(row, row, r, row, r)
        ws['GM{}'.format(
            row)] = '=IFERROR(ROUND(IF(FP{}="","",IF(70+30*GB{}/$GB${}<20,20,70+30*GB{}/$GB${})),2),"")'.format(row, row, r, row, r)
        ws['GN{}'.format(
            row)] = '=IFERROR(ROUND(IF(FQ{}="","",IF(70+30*GC{}/$GC${}<20,20,70+30*GC{}/$GC${})),2),"")'.format(row, row, r, row, r)
        ws['GO{}'.format(
            row)] = '=IFERROR(ROUND(IF(FR{}="","",IF(70+30*GD{}/$GD${}<20,20,70+30*GD{}/$GD${})),2),"")'.format(row, row, r, row, r)
        ws['GP{}'.format(
            row)] = '=IFERROR(ROUND(IF(FS{}="","",IF(70+30*GE{}/$GE${}<20,20,70+30*GE{}/$GE${})),2),"")'.format(row, row, r, row, r)
        ws['GQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(FT{}="","",IF(70+30*GF{}/$GF${}<20,20,70+30*GF{}/$GF${})),2),"")'.format(row, row, r, row, r)
        ws['GR{}'.format(
            row)] = '=IFERROR(ROUND(IF(FU{}="","",IF(70+30*GG{}/$GG${}<20,20,70+30*GG{}/$GG${})),2),"")'.format(row, row, r, row, r)
        ws['GS{}'.format(
            row)] = '=IFERROR(ROUND(IF(FV{}="","",IF(70+30*GH{}/$GH${}<20,20,70+30*GH{}/$GH${})),2),"")'.format(row, row, r, row, r)
        ws['GT{}'.format(
            row)] = '=IFERROR(ROUND(IF(FW{}="","",IF(70+30*GI{}/$GI${}<20,20,70+30*GI{}/$GI${})),2),"")'.format(row, row, r, row, r)

        ws['GU{}'.format(row)] = '=IF(SUM(GJ{}:GT{})=0,"",SUM(GJ{}:GT{}))'.format(
            row, row, row, row)
        ws['GV{}'.format(row)] = '=IF(GU{}="","",RANK(GU{},$GU$2:$GU${}))'.format(
            row, row, q)
        ws['GW{}'.format(
            row)] = '=IF(GV{}="","",COUNTIFS($FL$2:$FL${},FL{},$GV$2:$GV${},"<"&GV{})+1)'.format(row, q, row, q, row)

    # TAMBAHAN
        ws['GX{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,GJ{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,GJ{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,GJ{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,GJ{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,GJ{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,GJ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GY{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,GK{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,GK{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,GK{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,GK{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,GK{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,GK{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GZ{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,GL{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,GL{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,GL{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,GL{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,GL{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,GL{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HA{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,GM{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,GM{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,GM{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,GM{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,GM{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,GM{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HB{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,GN{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,GN{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,GN{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,GN{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,GN{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,GN{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HC{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,GO{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,GO{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,GO{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,GO{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,GO{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,GO{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HD{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,GP{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,GP{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,GP{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,GP{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,GP{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,GP{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HE{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,GQ{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,GQ{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,GQ{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,GQ{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,GQ{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,GQ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HF{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,GR{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,GR{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,GR{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,GR{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,GR{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,GR{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HG{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,GS{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,GS{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,GS{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,GS{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,GS{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,GS{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HH{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,GT{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,GT{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,GT{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,GT{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,GT{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,GT{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [5]
    ws['HJ1'] = 'NAMA SISWA_A'
    ws['HK1'] = 'NOMOR NF_A'
    ws['HL1'] = 'KELAS_A'
    ws['HM1'] = 'NAMA SEKOLAH_A'
    ws['HN1'] = 'LOKASI_A'

    ws['HO1'] = 'MAW_A'
    ws['HP1'] = 'MAP_A'
    ws['HQ1'] = 'IND_A'
    ws['HR1'] = 'ENG_A'
    ws['HS1'] = 'SEJ_A'
    ws['HT1'] = 'GEO_A'
    ws['HU1'] = 'EKO_A'
    ws['HV1'] = 'SOS_A'
    ws['HW1'] = 'FIS_A'
    ws['HX1'] = 'KIM_A'
    ws['HY1'] = 'BIO_A'
    ws['HZ1'] = 'JML_A'

    ws['IA1'] = 'Z_MAW_A'
    ws['IB1'] = 'Z_MAP_A'
    ws['IC1'] = 'Z_IND_A'
    ws['ID1'] = 'Z_ENG_A'
    ws['IE1'] = 'Z_SEJ_A'
    ws['IF1'] = 'Z_GEO_A'
    ws['IG1'] = 'Z_EKO_A'
    ws['IH1'] = 'Z_SOS_A'
    ws['II1'] = 'Z_FIS_A'
    ws['IJ1'] = 'Z_KIM_A'
    ws['IK1'] = 'Z_BIO_A'

    ws['IL1'] = 'S_MAW_A'
    ws['IM1'] = 'S_MAP_A'
    ws['IN1'] = 'S_IND_A'
    ws['IO1'] = 'S_ENG_A'
    ws['IP1'] = 'S_SEJ_A'
    ws['IQ1'] = 'S_GEO_A'
    ws['IR1'] = 'S_EKO_A'
    ws['IS1'] = 'S_SOS_A'
    ws['IT1'] = 'S_FIS_A'
    ws['IU1'] = 'S_KIM_A'
    ws['IV1'] = 'S_BIO_A'
    ws['IW1'] = 'S_JML_A'

    ws['IX1'] = 'RANK NAS._A'
    ws['IY1'] = 'RANK LOK._A'

    ws['IA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ID1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['II1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['IY1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['HJ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HK1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HL1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HM1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HN1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HO1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HP1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HQ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HR1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HS1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HT1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HU1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HV1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HW1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HX1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HY1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['HZ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IA1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IB1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IC1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['ID1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IE1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IF1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IG1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IH1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['II1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IJ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IK1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IL1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IM1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IN1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IO1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IP1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IQ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IR1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IS1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IT1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IU1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IV1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IW1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IX1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['IY1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

    # tambahan
    ws['IZ1'] = 'MAW_20_A'
    ws['JA1'] = 'MAP_20_A'
    ws['JB1'] = 'IND_20_A'
    ws['JC1'] = 'ENG_20_A'
    ws['JD1'] = 'SEJ_20_A'
    ws['JE1'] = 'GEO_20_A'
    ws['JF1'] = 'EKO_20_A'
    ws['JG1'] = 'SOS_20_A'
    ws['JH1'] = 'FIS_20_A'
    ws['JI1'] = 'KIM_20_A'
    ws['JJ1'] = 'BIO_20_A'

    ws['IZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['JJ1'].font = Font(bold=False, name='Calibri', size=11)

    ws['IZ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JA1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JB1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JC1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JD1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JE1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JF1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JG1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JH1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JI1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['JJ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

    for row in range(2, q+1):
        ws['HJ{}'.format(row)] = '=B{}'.format(row)
        ws['HK{}'.format(row)] = '=C{}'.format(row, row)
        ws['HL{}'.format(row)] = '=D{}'.format(row, row)
        ws['HM{}'.format(row)] = '=E{}'.format(row, row)
        ws['HN{}'.format(row)] = '=F{}'.format(row, row)

        ws['HO{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
        ws['HP{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
        ws['HQ{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
        ws['HR{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
        ws['HS{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
        ws['HT{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
        ws['HU{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
        ws['HV{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
        ws['HW{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
        ws['HX{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
        ws['HY{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
        ws['HZ{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
        ws['IA{}'.format(
            row)] = '=IFERROR(ROUND(IF(HO{}="","",(HO{}-HO${})/HO${}),2),"")'.format(row, row, r, s)
        ws['IB{}'.format(
            row)] = '=IFERROR(ROUND(IF(HP{}="","",(HP{}-HP${})/HP${}),2),"")'.format(row, row, r, s)
        ws['IC{}'.format(
            row)] = '=IFERROR(ROUND(IF(HQ{}="","",(HQ{}-HQ${})/HQ${}),2),"")'.format(row, row, r, s)
        ws['ID{}'.format(
            row)] = '=IFERROR(ROUND(IF(HR{}="","",(HR{}-HR${})/HR${}),2),"")'.format(row, row, r, s)
        ws['IE{}'.format(
            row)] = '=IFERROR(ROUND(IF(HS{}="","",(HS{}-HS${})/HS${}),2),"")'.format(row, row, r, s)
        ws['IF{}'.format(
            row)] = '=IFERROR(ROUND(IF(HT{}="","",(HT{}-HT${})/HT${}),2),"")'.format(row, row, r, s)
        ws['IG{}'.format(
            row)] = '=IFERROR(ROUND(IF(HU{}="","",(HU{}-HU${})/HU${}),2),"")'.format(row, row, r, s)
        ws['IH{}'.format(
            row)] = '=IFERROR(ROUND(IF(HV{}="","",(HV{}-HV${})/HV${}),2),"")'.format(row, row, r, s)
        ws['II{}'.format(
            row)] = '=IFERROR(ROUND(IF(HW{}="","",(HW{}-HW${})/HW${}),2),"")'.format(row, row, r, s)
        ws['IJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(HX{}="","",(HX{}-HX${})/HX${}),2),"")'.format(row, row, r, s)
        ws['IK{}'.format(
            row)] = '=IFERROR(ROUND(IF(HY{}="","",(HY{}-HY${})/HY${}),2),"")'.format(row, row, r, s)

        ws['IL{}'.format(
            row)] = '=IFERROR(ROUND(IF(HO{}="","",IF(70+30*IA{}/$IA${}<20,20,70+30*IA{}/$IA${})),2),"")'.format(row, row, r, row, r)
        ws['IM{}'.format(
            row)] = '=IFERROR(ROUND(IF(HP{}="","",IF(70+30*IB{}/$IB${}<20,20,70+30*IB{}/$IB${})),2),"")'.format(row, row, r, row, r)
        ws['IN{}'.format(
            row)] = '=IFERROR(ROUND(IF(HQ{}="","",IF(70+30*IC{}/$IC${}<20,20,70+30*IC{}/$IC${})),2),"")'.format(row, row, r, row, r)
        ws['IO{}'.format(
            row)] = '=IFERROR(ROUND(IF(HR{}="","",IF(70+30*ID{}/$ID${}<20,20,70+30*ID{}/$ID${})),2),"")'.format(row, row, r, row, r)
        ws['IP{}'.format(
            row)] = '=IFERROR(ROUND(IF(HS{}="","",IF(70+30*IE{}/$IE${}<20,20,70+30*IE{}/$IE${})),2),"")'.format(row, row, r, row, r)
        ws['IQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(HT{}="","",IF(70+30*IF{}/$IF${}<20,20,70+30*IF{}/$IF${})),2),"")'.format(row, row, r, row, r)
        ws['IR{}'.format(
            row)] = '=IFERROR(ROUND(IF(HU{}="","",IF(70+30*IG{}/$IG${}<20,20,70+30*IG{}/$IG${})),2),"")'.format(row, row, r, row, r)
        ws['IS{}'.format(
            row)] = '=IFERROR(ROUND(IF(HV{}="","",IF(70+30*IH{}/$IH${}<20,20,70+30*IH{}/$IH${})),2),"")'.format(row, row, r, row, r)
        ws['IT{}'.format(
            row)] = '=IFERROR(ROUND(IF(HW{}="","",IF(70+30*II{}/$II${}<20,20,70+30*II{}/$II${})),2),"")'.format(row, row, r, row, r)
        ws['IU{}'.format(
            row)] = '=IFERROR(ROUND(IF(HX{}="","",IF(70+30*IJ{}/$IJ${}<20,20,70+30*IJ{}/$IJ${})),2),"")'.format(row, row, r, row, r)
        ws['IV{}'.format(
            row)] = '=IFERROR(ROUND(IF(FW{}="","",IF(70+30*IK{}/$IK${}<20,20,70+30*IK{}/$IK${})),2),"")'.format(row, row, r, row, r)

        ws['IW{}'.format(row)] = '=IF(SUM(IL{}:IV{})=0,"",SUM(IL{}:IV{}))'.format(
            row, row, row, row)
        ws['IX{}'.format(row)] = '=IF(IW{}="","",RANK(IW{},$IW$2:$IW${}))'.format(
            row, row, q)
        ws['IY{}'.format(
            row)] = '=IF(IX{}="","",COUNTIFS($HN$2:$HN${},HN{},$IX$2:$IX${},"<"&IX{})+1)'.format(row, q, row, q, row)

    # TAMBAHAN
        ws['IZ{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,IL{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,IL{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,IL{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,IL{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,IL{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,IL{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JA{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,IM{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,IM{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,IM{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,IM{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,IM{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,IM{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JB{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,IN{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,IN{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,IN{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,IN{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,IN{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,IN{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JC{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,IO{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,IO{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,IO{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,IO{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,IO{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,IO{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JD{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,IP{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,IP{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,IP{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,IP{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,IP{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,IP{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JE{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,IQ{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,IQ{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,IQ{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,IQ{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,IQ{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,IQ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JF{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,IR{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,IR{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,IR{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,IR{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,IR{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,IR{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JG{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,IS{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,IS{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,IS{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,IS{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,IS{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,IS{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JH{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,IT{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,IT{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,IT{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,IT{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,IT{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,IT{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JI{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,IU{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,IU{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,IU{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,IU{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,IU{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,IU{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['JJ{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,IV{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,IV{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,IV{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,IV{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,IV{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,IV{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score
    ws['JL1'] = 'NAMA SISWA'
    ws['JM1'] = 'NOMOR NF'
    ws['JN1'] = 'KELAS'
    ws['JO1'] = 'NAMA SEKOLAH'
    ws['JP1'] = 'LOKASI'

    ws['JQ1'] = 'MAW'
    ws['JR1'] = 'MAP'
    ws['JS1'] = 'IND'
    ws['JT1'] = 'ENG'
    ws['JU1'] = 'SEJ'
    ws['JV1'] = 'GEO'
    ws['JW1'] = 'EKO'
    ws['JX1'] = 'SOS'
    ws['JY1'] = 'FIS'
    ws['JZ1'] = 'KIM'
    ws['KA1'] = 'BIO'
    ws['KB1'] = 'JML'

    ws['KC1'] = 'Z_MAW'
    ws['KD1'] = 'Z_MAP'
    ws['KE1'] = 'Z_IND'
    ws['KF1'] = 'Z_ENG'
    ws['KG1'] = 'Z_SEJ'
    ws['KH1'] = 'Z_GEO'
    ws['KI1'] = 'Z_EKO'
    ws['KJ1'] = 'Z_SOS'
    ws['KK1'] = 'Z_FIS'
    ws['KL1'] = 'Z_KIM'
    ws['KM1'] = 'Z_BIO'

    ws['KN1'] = 'S_MAW'
    ws['KO1'] = 'S_MAP'
    ws['KP1'] = 'S_IND'
    ws['KQ1'] = 'S_ENG'
    ws['KR1'] = 'S_SEJ'
    ws['KS1'] = 'S_GEO'
    ws['KT1'] = 'S_EKO'
    ws['KU1'] = 'S_SOS'
    ws['KV1'] = 'S_FIS'
    ws['KW1'] = 'S_KIM'
    ws['KX1'] = 'S_BIO'
    ws['KY1'] = 'S_JML'

    ws['KZ1'] = 'RANK NAS.'
    ws['LA1'] = 'RANK LOK.'

    ws['KC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['KZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LA1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['JL1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JM1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JN1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JO1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JP1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JQ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JR1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JS1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JT1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JU1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JV1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JW1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JX1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JY1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['JZ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KA1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KB1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KC1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KD1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KE1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KF1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KG1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KH1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KI1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KJ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KK1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KL1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KM1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KN1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KO1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KP1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KQ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KR1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KS1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KT1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KU1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KV1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KW1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KX1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KY1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['KZ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LA1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

    # tambahan
    ws['LB1'] = 'MAW_20'
    ws['LC1'] = 'MAP_20'
    ws['LD1'] = 'IND_20'
    ws['LE1'] = 'ENG_20'
    ws['LF1'] = 'SEJ_20'
    ws['LG1'] = 'GEO_20'
    ws['LH1'] = 'EKO_20'
    ws['LI1'] = 'SOS_20'
    ws['LJ1'] = 'FIS_20'
    ws['LK1'] = 'KIM_20'
    ws['LL1'] = 'BIO_20'

    ws['LB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['LL1'].font = Font(bold=False, name='Calibri', size=11)

    ws['LB1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LC1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LD1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LE1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LF1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LG1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LH1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LI1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LJ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LK1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['LL1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

    for row in range(2, q+1):
        ws['JL{}'.format(row)] = '=B{}'.format(row)
        ws['JM{}'.format(row)] = '=C{}'.format(row, row)
        ws['JN{}'.format(row)] = '=D{}'.format(row, row)
        ws['JO{}'.format(row)] = '=E{}'.format(row, row)
        ws['JP{}'.format(row)] = '=F{}'.format(row, row)

        ws['JQ{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
        ws['JR{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
        ws['JS{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
        ws['JT{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
        ws['JU{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
        ws['JV{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
        ws['JW{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
        ws['JX{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
        ws['JY{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
        ws['JZ{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
        ws['KA{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
        ws['KB{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
        ws['KC{}'.format(
            row)] = '=IFERROR(ROUND(IF(JQ{}="","",(JQ{}-JQ${})/JQ${}),2),"")'.format(row, row, r, s)
        ws['KD{}'.format(
            row)] = '=IFERROR(ROUND(IF(JR{}="","",(JR{}-JR${})/JR${}),2),"")'.format(row, row, r, s)
        ws['KE{}'.format(
            row)] = '=IFERROR(ROUND(IF(JS{}="","",(JS{}-JS${})/JS${}),2),"")'.format(row, row, r, s)
        ws['KF{}'.format(
            row)] = '=IFERROR(ROUND(IF(JT{}="","",(JT{}-JT${})/JT${}),2),"")'.format(row, row, r, s)
        ws['KG{}'.format(
            row)] = '=IFERROR(ROUND(IF(JU{}="","",(JU{}-JU${})/JU${}),2),"")'.format(row, row, r, s)
        ws['KH{}'.format(
            row)] = '=IFERROR(ROUND(IF(JV{}="","",(JV{}-JV${})/JV${}),2),"")'.format(row, row, r, s)
        ws['KI{}'.format(
            row)] = '=IFERROR(ROUND(IF(JW{}="","",(JW{}-JW${})/JW${}),2),"")'.format(row, row, r, s)
        ws['KJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(JX{}="","",(JX{}-JX${})/JX${}),2),"")'.format(row, row, r, s)
        ws['KK{}'.format(
            row)] = '=IFERROR(ROUND(IF(JY{}="","",(JY{}-JY${})/JY${}),2),"")'.format(row, row, r, s)
        ws['KL{}'.format(
            row)] = '=IFERROR(ROUND(IF(JZ{}="","",(JZ{}-JZ${})/JZ${}),2),"")'.format(row, row, r, s)
        ws['KM{}'.format(
            row)] = '=IFERROR(ROUND(IF(KA{}="","",(KA{}-KA${})/KA${}),2),"")'.format(row, row, r, s)

        ws['KN{}'.format(
            row)] = '=IFERROR(ROUND(IF(JQ{}="","",IF(70+30*KC{}/$KC${}<20,20,70+30*KC{}/$KC${})),2),"")'.format(row, row, r, row, r)
        ws['KO{}'.format(
            row)] = '=IFERROR(ROUND(IF(JR{}="","",IF(70+30*KD{}/$KD${}<20,20,70+30*KD{}/$KD${})),2),"")'.format(row, row, r, row, r)
        ws['KP{}'.format(
            row)] = '=IFERROR(ROUND(IF(JS{}="","",IF(70+30*KE{}/$KE${}<20,20,70+30*KE{}/$KE${})),2),"")'.format(row, row, r, row, r)
        ws['KQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(JT{}="","",IF(70+30*KF{}/$KF${}<20,20,70+30*KF{}/$KF${})),2),"")'.format(row, row, r, row, r)
        ws['KR{}'.format(
            row)] = '=IFERROR(ROUND(IF(JU{}="","",IF(70+30*KG{}/$KG${}<20,20,70+30*KG{}/$KG${})),2),"")'.format(row, row, r, row, r)
        ws['KS{}'.format(
            row)] = '=IFERROR(ROUND(IF(JV{}="","",IF(70+30*KH{}/$KH${}<20,20,70+30*KH{}/$KH${})),2),"")'.format(row, row, r, row, r)
        ws['KT{}'.format(
            row)] = '=IFERROR(ROUND(IF(JW{}="","",IF(70+30*KI{}/$KI${}<20,20,70+30*KI{}/$KI${})),2),"")'.format(row, row, r, row, r)
        ws['KU{}'.format(
            row)] = '=IFERROR(ROUND(IF(JX{}="","",IF(70+30*KJ{}/$KJ${}<20,20,70+30*KJ{}/$KJ${})),2),"")'.format(row, row, r, row, r)
        ws['KV{}'.format(
            row)] = '=IFERROR(ROUND(IF(JY{}="","",IF(70+30*KK{}/$KK${}<20,20,70+30*KK{}/$KK${})),2),"")'.format(row, row, r, row, r)
        ws['KW{}'.format(
            row)] = '=IFERROR(ROUND(IF(JZ{}="","",IF(70+30*KL{}/$KL${}<20,20,70+30*KL{}/$KL${})),2),"")'.format(row, row, r, row, r)
        ws['KX{}'.format(
            row)] = '=IFERROR(ROUND(IF(FW{}="","",IF(70+30*KM{}/$KM${}<20,20,70+30*KM{}/$KM${})),2),"")'.format(row, row, r, row, r)

        ws['KY{}'.format(row)] = '=IF(SUM(KN{}:KX{})=0,"",SUM(KN{}:KX{}))'.format(
            row, row, row, row)
        ws['KZ{}'.format(row)] = '=IF(KY{}="","",RANK(KY{},$KY$2:$KY${}))'.format(
            row, row, q)
        ws['LA{}'.format(
            row)] = '=IF(KZ{}="","",COUNTIFS($JP$2:$JP${},JP{},$KZ$2:$KZ${},"<"&KZ{})+1)'.format(row, q, row, q, row)

    # TAMBAHAN
        ws['LB{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,KN{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,KN{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,KN{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,KN{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,KN{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,KN{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LC{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,KO{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,KO{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,KO{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,KO{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,KO{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,KO{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LD{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,KP{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,KP{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,KP{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,KP{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,KP{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,KP{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LE{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,KQ{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,KQ{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,KQ{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,KQ{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,KQ{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,KQ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LF{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,KR{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,KR{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,KR{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,KR{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,KR{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,KR{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LG{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,KS{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,KS{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,KS{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,KS{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,KS{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,KS{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LH{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,KT{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,KT{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,KT{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,KT{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,KT{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,KT{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LI{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,KU{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,KU{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,KU{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,KU{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,KU{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,KU{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LJ{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,KV{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,KV{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,KV{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,KV{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,KV{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,KV{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LK{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,KW{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,KW{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,KW{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,KW{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,KW{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,KW{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['LL{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,KX{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,KX{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,KX{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,KX{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,KX{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,KX{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
    kelas = KELAS.lower().replace(" ", "")
    semester = SEMESTER.lower()
    tahun = TAHUN.replace("-", "")
    penilaian = PENILAIAN.lower()
    kurikulum = KURIKULUM.lower()

    path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

    # Simpan file ke direktori temporer
    temp_dir = tempfile.gettempdir()
    file_path = temp_dir + '/' + path_file
    wb.save(file_path)

    st.success(
        "File siap diunduh!")

    # Tombol unduh file
    with open(file_path, "rb") as f:
        bytes_data = f.read()
    st.download_button(label="Unduh File", data=bytes_data,
                       file_name=path_file)

    st.warning(
        "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
